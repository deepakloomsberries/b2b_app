import csv
import io
import secrets
import sqlite3
import json
import os
import random
import time
import urllib.request
from datetime import datetime, timedelta

from dotenv import load_dotenv
from flask import (
    Flask,
    flash,
    redirect,
    Response,
    render_template,
    jsonify,
    request,
    send_file,
    session,
    url_for,
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import load_workbook

from db import get_db, get_setting, init_db, now_iso, set_setting
import google_sheets

load_dotenv()

app = Flask(__name__)
_secret_key = os.getenv("FLASK_SECRET_KEY")
if not _secret_key:
    import secrets as _secrets
    _secret_key = _secrets.token_hex(32)
    print(
        "[WARNING] FLASK_SECRET_KEY not set. Sessions will not persist across restarts. "
        "Set FLASK_SECRET_KEY in your .env file.",
        flush=True,
    )
app.secret_key = _secret_key
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_REFRESH_EACH_REQUEST"] = True
app.config["SESSION_PERMANENT"] = True
app.permanent_session_lifetime = timedelta(days=30)

_db_initialized = False
_catalog_cache = {"timestamp": 0, "data": None}
_CATALOG_CACHE_TTL = 120


def _seed_default_admin():
    """Create a default admin user on first run if none exists."""
    import secrets as _secrets
    with get_db() as conn:
        count = conn.execute(
            "SELECT COUNT(*) FROM users WHERE role = 'admin'"
        ).fetchone()[0]
        if count > 0:
            return
        username = os.getenv("ADMIN_USERNAME", "admin")
        password = os.getenv("ADMIN_PASSWORD", "")
        if not password:
            password = _secrets.token_urlsafe(12)
            print(
                f"[SETUP] No ADMIN_PASSWORD set. Generated admin password: {password}",
                flush=True,
            )
            print(
                "[SETUP] Set ADMIN_USERNAME and ADMIN_PASSWORD in your .env to use fixed credentials.",
                flush=True,
            )
        password_hash = generate_password_hash(password)
        conn.execute(
            "INSERT INTO users (name, role, password, created_at) VALUES (?, 'admin', ?, ?)",
            (username, password_hash, now_iso()),
        )
        conn.commit()
        print(f"[SETUP] Default admin user '{username}' created.", flush=True)


@app.before_request
def setup_database():
    global _db_initialized
    if not _db_initialized:
        init_db()
        _seed_default_admin()
        _db_initialized = True


@app.before_request
def require_login():
    open_endpoints = {"login", "static", "service_worker"}
    if request.endpoint in open_endpoints or request.endpoint is None:
        return None
    if not session.get("user"):
        return redirect(url_for("login"))
    session.permanent = True
    session.modified = True
    return None


CSRF_SESSION_KEY = "_csrf_token"


def get_csrf_token():
    token = session.get(CSRF_SESSION_KEY)
    if not token:
        token = secrets.token_urlsafe(32)
        session[CSRF_SESSION_KEY] = token
    return token


app.jinja_env.globals["csrf_token"] = get_csrf_token


@app.before_request
def enforce_csrf():
    if request.method not in {"POST", "PUT", "PATCH", "DELETE"}:
        return None
    expected = session.get(CSRF_SESSION_KEY)
    header_token = request.headers.get("X-CSRFToken")
    submitted = request.form.get("csrf_token") or header_token
    if not expected or not submitted or not secrets.compare_digest(submitted, expected):
        # fetch() follows redirects and lands on a 200, so a caller that
        # authenticates via the header (JS) must get a real error status -
        # otherwise `response.ok` looks like success for a request that
        # never went through.
        if header_token is not None:
            return jsonify({"error": "csrf_failed"}), 400
        flash("Your session expired. Please try again.", "error")
        return redirect(request.referrer or url_for("home"))
    return None


def require_admin():
    user = session.get("user")
    if not user or normalize_user_role(user.get("role")) != "admin":
        flash("Admin access required.", "error")
        return redirect(url_for("home"))
    return None


def get_orders_view_param():
    view = (request.args.get("view") or request.form.get("view") or "").strip().lower()
    return view if view in {"live", "history"} else "history"


def parse_iso_date(value):
    if not value:
        return None
    try:
        cleaned = str(value).strip()
        if cleaned.endswith("Z"):
            cleaned = cleaned[:-1]
        return datetime.fromisoformat(cleaned)
    except ValueError:
        return None


def parse_float(value, default=0.0):
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return default


def parse_int(value):
    if value is None:
        return None
    if isinstance(value, int):
        return value
    try:
        return int(str(value).strip())
    except ValueError:
        return None


def normalize_user_role(role):
    normalized = str(role or "user").strip().lower()
    return normalized if normalized in {"user", "warehouse", "admin"} else "user"


LOGIN_MAX_ATTEMPTS = 5
LOGIN_LOCKOUT_MINUTES = 15


def check_login_lockout(conn, identifier):
    row = conn.execute(
        "SELECT locked_until FROM login_attempts WHERE identifier = ?", (identifier,)
    ).fetchone()
    if not row or not row["locked_until"]:
        return None
    locked_until = parse_iso_date(row["locked_until"])
    if locked_until and locked_until > datetime.utcnow():
        return locked_until
    return None


def register_failed_login(conn, identifier):
    row = conn.execute(
        "SELECT fail_count FROM login_attempts WHERE identifier = ?", (identifier,)
    ).fetchone()
    fail_count = (row["fail_count"] if row else 0) + 1
    locked_until = None
    if fail_count >= LOGIN_MAX_ATTEMPTS:
        locked_until = (
            datetime.utcnow() + timedelta(minutes=LOGIN_LOCKOUT_MINUTES)
        ).isoformat()
        fail_count = 0
    conn.execute(
        "INSERT INTO login_attempts (identifier, fail_count, locked_until, last_attempt_at) "
        "VALUES (?, ?, ?, ?) "
        "ON CONFLICT(identifier) DO UPDATE SET fail_count = excluded.fail_count, "
        "locked_until = excluded.locked_until, last_attempt_at = excluded.last_attempt_at",
        (identifier, fail_count, locked_until, now_iso()),
    )
    conn.commit()


def clear_login_attempts(conn, identifier):
    conn.execute("DELETE FROM login_attempts WHERE identifier = ?", (identifier,))
    conn.commit()


def require_warehouse():
    user = session.get("user")
    if not user or normalize_user_role(user.get("role")) not in {"warehouse", "admin"}:
        flash("Warehouse or admin access required.", "error")
        return redirect(url_for("orders_list"))
    return None


def sanitize_csv_cell(value):
    """Prefix a leading apostrophe on values that could be read as spreadsheet
    formulas (=, +, -, @) so downloaded/exported CSVs open safely in Excel/Sheets."""
    text = "" if value is None else str(value)
    if text[:1] in ("=", "+", "-", "@"):
        return "'" + text
    return text


def get_salesman_price_overrides(conn, user_id):
    """Return {sku: price} of this salesman's custom prices, or {} if none."""
    if not user_id:
        return {}
    rows = conn.execute(
        "SELECT sku, price FROM salesman_prices WHERE user_id = ?", (user_id,)
    ).fetchall()
    return {row["sku"]: row["price"] for row in rows}


def resolve_product_base_price(sku, product_row, overrides):
    """Effective unit price for this salesman: their override if set, else the
    product's standard price (price, falling back to price_credit)."""
    if sku in overrides:
        return overrides[sku]
    return (
        product_row["price"]
        if product_row["price"] is not None
        else product_row["price_credit"]
    )


def get_min_price_percent():
    try:
        value = float(get_setting("min_price_percent", "100"))
    except (TypeError, ValueError):
        value = 100.0
    return max(0.0, min(100.0, value))


def get_vat_rate():
    """VAT rate as a decimal fraction (e.g. 0.15), from the vat_rate setting (percent)."""
    try:
        value = float(get_setting("vat_rate", "15"))
    except (TypeError, ValueError):
        value = 15.0
    return max(0.0, min(100.0, value)) / 100.0


def static_url(filename):
    """URL for a static file with a cache-busting ?v=<mtime> query string, so
    browsers fetch new CSS/JS right after a deploy instead of serving a stale
    cached copy until the user does a hard refresh."""
    file_path = os.path.join(app.static_folder, filename)
    try:
        version = int(os.path.getmtime(file_path))
    except OSError:
        version = 0
    return url_for("static", filename=filename, v=version)


app.jinja_env.globals["static_url"] = static_url


def get_credit_warning(customer):
    """Return a warning dict if the customer's outstanding balance exceeds the
    configured credit limit, or None if the limit is disabled (0) or not exceeded."""
    credit_limit = parse_float(get_setting("credit_limit", "0"))
    outstanding_balance = parse_float(customer["outstanding_balance"]) if customer else 0.0
    if credit_limit <= 0 or outstanding_balance <= credit_limit:
        return None
    return {
        "outstanding_balance": outstanding_balance,
        "credit_limit": credit_limit,
    }


def format_datetime(value):
    if not value:
        return ""
    try:
        cleaned = str(value).strip()
        if cleaned.endswith("Z"):
            cleaned = cleaned[:-1]
        dt = datetime.fromisoformat(cleaned)
        return dt.strftime("%d %b %Y, %I:%M %p")
    except ValueError:
        return str(value)


def format_date(value):
    if not value:
        return ""
    try:
        cleaned = str(value).strip()
        if cleaned.endswith("Z"):
            cleaned = cleaned[:-1]
        dt = datetime.fromisoformat(cleaned)
        return dt.strftime("%d-%B-%Y")
    except ValueError:
        return str(value)


def has_recent_duplicate_order(conn, customer_id, submitter_name, valid_items, cutoff_iso):
    item_totals = {sku: qty for sku, qty, *_ in valid_items}
    if not item_totals:
        return None
    candidates = conn.execute(
        "SELECT id, order_number FROM orders "
        "WHERE customer_id = ? AND COALESCE(submitted_by, '') = ? AND created_at >= ? "
        "ORDER BY created_at DESC",
        (customer_id, submitter_name, cutoff_iso),
    ).fetchall()
    for candidate in candidates:
        candidate_items = conn.execute(
            "SELECT sku, qty FROM order_items WHERE order_id = ?",
            (candidate["id"],),
        ).fetchall()
        candidate_totals = {}
        for row in candidate_items:
            candidate_totals[row["sku"]] = candidate_totals.get(row["sku"], 0) + int(
                row["qty"] or 0
            )
        if candidate_totals == item_totals:
            return candidate["order_number"]
    return None


app.add_template_filter(format_datetime, "format_dt")
app.add_template_filter(format_date, "format_date")


def log_audit(conn, action, entity, details=None):
    actor = session.get("user", {}).get("name", "Unknown")
    payload = json.dumps(details or {}, ensure_ascii=False)
    conn.execute(
        "INSERT INTO audit_logs (actor, action, entity, details, created_at) "
        "VALUES (?, ?, ?, ?, ?)",
        (actor, action, entity, payload, now_iso()),
    )


def get_active_notifications(conn, user_role, user_name, limit=20):
    return conn.execute(
        """
        SELECT n.id,
               n.title,
               n.message,
               n.priority,
               n.action_url,
               n.created_at,
               COALESCE(n.expires_at, '') AS expires_at,
               n.audience_role,
               CASE WHEN nr.id IS NULL THEN 0 ELSE 1 END AS is_read
        FROM notifications n
        LEFT JOIN notification_reads nr
            ON nr.notification_id = n.id
            AND nr.reader_name = ?
        WHERE (n.audience_role = 'all' OR n.audience_role = ?)
          AND (n.expires_at IS NULL OR n.expires_at = '' OR n.expires_at > ?)
        ORDER BY n.created_at DESC
        LIMIT ?
        """,
        (user_name, user_role, now_iso(), limit),
    ).fetchall()


def parse_datetime_local(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(value).isoformat()
    except ValueError:
        return None


def create_notification(
    conn,
    title,
    message,
    audience_role="all",
    priority="info",
    action_url=None,
    expires_at=None,
):
    conn.execute(
        "INSERT INTO notifications (title, message, audience_role, priority, action_url, expires_at, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        (
            title,
            message,
            audience_role,
            priority,
            action_url,
            expires_at,
            now_iso(),
        ),
    )


def load_catalog_rows(file_storage):
    filename = (file_storage.filename or "").lower()
    ext = os.path.splitext(filename)[1]
    raw_bytes = file_storage.read()
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return None, "The uploaded Excel file is empty."
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        data_rows = []
        for row in rows[1:]:
            row_dict = {}
            for header, cell in zip(headers, row):
                if header:
                    row_dict[header] = "" if cell is None else str(cell)
            if any(value for value in row_dict.values()):
                data_rows.append(row_dict)
        return data_rows, None
    if ext == ".csv":
        try:
            decoded = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            decoded = raw_bytes.decode("latin-1", errors="replace")
        reader = csv.DictReader(decoded.splitlines())
        return list(reader), None
    return None, "Please upload a CSV or Excel (.xlsx) file."


def sync_reserved_stock(conn):
    for attempt in range(3):
        try:
            conn.execute("BEGIN IMMEDIATE")
            conn.execute("DELETE FROM reserved_stock")
            conn.execute(
                "INSERT INTO reserved_stock (sku, reserved_qty) "
                "SELECT oi.sku, SUM(oi.qty) "
                "FROM order_items oi "
                "JOIN orders o ON o.id = oi.order_id "
                "WHERE o.order_status IN ('submitted', 'accepted', 'packed', 'shipped') "
                "GROUP BY oi.sku"
            )
            conn.commit()
            return
        except sqlite3.OperationalError as error:
            conn.rollback()
            if "locked" not in str(error).lower() or attempt == 2:
                raise
            time.sleep(0.2 * (attempt + 1))


def ensure_reserved_stock(conn):
    existing = conn.execute(
        "SELECT 1 FROM reserved_stock LIMIT 1"
    ).fetchone()
    if existing is None:
        sync_reserved_stock(conn)


def adjust_reserved_stock(conn, order_id, multiplier):
    items = conn.execute(
        "SELECT sku, qty FROM order_items WHERE order_id = ?",
        (order_id,),
    ).fetchall()
    for item in items:
        conn.execute(
            "INSERT INTO reserved_stock (sku, reserved_qty) VALUES (?, ?) "
            "ON CONFLICT(sku) DO UPDATE SET reserved_qty = MAX(reserved_qty + ?, 0)",
            (item["sku"], item["qty"] * multiplier, item["qty"] * multiplier),
        )


def verify_password(conn, user, password):
    stored = user["password"]
    if stored.startswith("pbkdf2:") or stored.startswith("scrypt:"):
        return check_password_hash(stored, password)
    if stored == password:
        conn.execute(
            "UPDATE users SET password = ? WHERE id = ?",
            (generate_password_hash(password), user["id"]),
        )
        conn.commit()
        return True
    return False


def user_email_exists(email):
    with get_db() as conn:
        return (
            conn.execute("SELECT 1 FROM users WHERE email = ?", (email,)).fetchone()
            is not None
        )


def generate_order_number(conn):
    date_part = datetime.utcnow().strftime("%Y%m%d")
    while True:
        suffix = f"{random.randint(0, 9999):04d}"
        order_number = f"ORD-{date_part}-{suffix}"
        existing = conn.execute(
            "SELECT 1 FROM orders WHERE order_number = ?", (order_number,)
        ).fetchone()
        if not existing:
            return order_number


def get_cached_catalog_data():
    now = time.time()
    cached = _catalog_cache.get("data")
    if cached and now - _catalog_cache["timestamp"] < _CATALOG_CACHE_TTL:
        return cached

    reservation_mode = get_setting("reservation_mode", "on") == "on"
    with get_db() as conn:
        if reservation_mode:
            ensure_reserved_stock(conn)
        avail_expr = (
            "COALESCE(s.stock_qty, 0) - COALESCE(r.reserved_qty, 0)"
            if reservation_mode
            else "COALESCE(s.stock_qty, 0)"
        )
        products = conn.execute(
            "SELECT p.*, "
            "COALESCE(s.stock_qty, 0) AS stock_qty, "
            "COALESCE(r.reserved_qty, 0) AS reserved_qty, "
            f"{avail_expr} AS available_qty "
            "FROM products p "
            "LEFT JOIN stock s ON p.sku = s.sku "
            "LEFT JOIN reserved_stock r ON p.sku = r.sku "
            "WHERE p.is_active = 1 "
            "ORDER BY COALESCE(p.sort_order, 999999), p.title"
        ).fetchall()
        category_rows = conn.execute(
            "SELECT name, image_url FROM categories ORDER BY name"
        ).fetchall()
        subcategory_rows = conn.execute(
            "SELECT name, category_name, image_url FROM subcategories ORDER BY name"
        ).fetchall()
        sub_subcategory_rows = conn.execute(
            "SELECT name, subcategory_name, image_url FROM sub_subcategories ORDER BY name"
        ).fetchall()

    categories = sorted({p["category"] for p in products if p["category"]})
    subcategories = sorted({p["subcategory"] for p in products if p["subcategory"]})
    category_images = {row["name"]: row["image_url"] for row in category_rows}
    subcategory_images = {row["name"]: row["image_url"] for row in subcategory_rows}
    sub_subcategory_images = {
        row["name"]: row["image_url"] for row in sub_subcategory_rows
    }

    # Build hierarchy from active products so stale/unmapped taxonomy rows
    # do not appear in the catalog filters.
    subcategory_map = {}
    sub_subcategory_map = {}
    for product in products:
        category_name = product["category"] or ""
        subcategory_name = product["subcategory"] or ""
        sub_subcategory_name = product["sub_subcategory"] or ""

        if subcategory_name:
            subcategory_map.setdefault(category_name, set()).add(subcategory_name)
        if sub_subcategory_name:
            sub_subcategory_map.setdefault(subcategory_name, set()).add(
                sub_subcategory_name
            )

    data = {
        "products": products,
        "categories": categories,
        "subcategories": subcategories,
        "category_images": category_images,
        "subcategory_map": subcategory_map,
        "subcategory_images": subcategory_images,
        "sub_subcategory_map": sub_subcategory_map,
        "sub_subcategory_images": sub_subcategory_images,
    }
    _catalog_cache["timestamp"] = now
    _catalog_cache["data"] = data
    return data


@app.route("/")
def home():
    if not session.get("user"):
        return redirect(url_for("login"))
    query = request.args.get("q", "").strip()
    with get_db() as conn:
        if query:
            customers = conn.execute(
                "SELECT * FROM customers WHERE name LIKE ? ORDER BY name",
                (f"%{query}%",),
            ).fetchall()
        else:
            customers = conn.execute("SELECT * FROM customers ORDER BY name").fetchall()

        total_buyers = conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
        draft_orders = conn.execute("SELECT COUNT(*) FROM drafts").fetchone()[0]
        month_prefix = datetime.utcnow().strftime("%Y-%m")
        orders_this_month = conn.execute(
            "SELECT COUNT(*) FROM orders WHERE created_at LIKE ?",
            (f"{month_prefix}%",),
        ).fetchone()[0]
        revenue_row = conn.execute(
            "SELECT SUM(oi.price_snapshot * oi.qty) AS total_revenue "
            "FROM orders o JOIN order_items oi ON o.id = oi.order_id "
            "WHERE o.created_at LIKE ?",
            (f"{month_prefix}%",),
        ).fetchone()
        revenue_this_month = revenue_row["total_revenue"] or 0

        recent_activity = []
        orders = conn.execute(
            "SELECT o.id, o.order_number, o.created_at, c.name AS customer_name "
            "FROM orders o JOIN customers c ON o.customer_id = c.id "
            "ORDER BY o.created_at DESC LIMIT 5"
        ).fetchall()
        customers_recent = conn.execute(
            "SELECT id, name, created_at FROM customers ORDER BY created_at DESC LIMIT 5"
        ).fetchall()
        for row in orders:
            recent_activity.append(
                {
                    "timestamp": row["created_at"],
                    "action": "Order placed",
                    "reference": row["order_number"],
                    "link": url_for("orders_list"),
                }
            )
        for row in customers_recent:
            recent_activity.append(
                {
                    "timestamp": row["created_at"],
                    "action": "Buyer created",
                    "reference": row["name"],
                    "link": url_for("customers_list"),
                }
            )
        recent_activity.sort(
            key=lambda item: item["timestamp"] or "", reverse=True
        )
        recent_activity = recent_activity[:5]

        beat_plan = conn.execute(
            "SELECT * FROM customers ORDER BY "
            "CASE priority_flag "
            "WHEN 'needs_visit' THEN 1 "
            "WHEN 'high_potential' THEN 2 "
            "WHEN 'on_hold' THEN 3 "
            "ELSE 4 END, "
            "CASE "
            "WHEN NULLIF(latitude, '') IS NULL OR NULLIF(longitude, '') IS NULL THEN 1 "
            "ELSE 0 END, "
            "CAST(NULLIF(latitude, '') AS REAL) ASC, "
            "CAST(NULLIF(longitude, '') AS REAL) ASC, "
            "COALESCE(last_visited_at, '') ASC "
            "LIMIT 5"
        ).fetchall()

    return render_template(
        "home_select_customer.html",
        customers=customers,
        query=query,
        total_buyers=total_buyers,
        draft_orders=draft_orders,
        orders_this_month=orders_this_month,
        revenue_this_month=revenue_this_month,
        recent_activity=recent_activity,
        beat_plan=beat_plan,
    )


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        normalized_username = username.lower()
        with get_db() as conn:
            user = conn.execute(
                "SELECT * FROM users WHERE LOWER(name) = ? OR LOWER(email) = ?",
                (normalized_username, normalized_username),
            ).fetchone()
            # Key the lockout on the resolved account name (not the typed
            # string) so alternating between a username and email for the
            # same account shares one attempt counter instead of doubling it.
            identifier = user["name"].lower() if user else normalized_username
            locked_until = check_login_lockout(conn, identifier)
            if locked_until:
                flash(
                    "Too many failed attempts. Try again after "
                    f"{locked_until.strftime('%H:%M UTC')}.",
                    "error",
                )
                return render_template("login.html")
            if user and verify_password(conn, user, password):
                clear_login_attempts(conn, identifier)
                session["user"] = {
                    "id": user["id"],
                    "name": user["name"],
                    "role": normalize_user_role(user["role"]),
                }
                session.permanent = True
                return redirect(url_for("home"))
            register_failed_login(conn, identifier)
        flash("Invalid credentials.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/service-worker.js")
def service_worker():
    return send_file(
        os.path.join(app.root_path, "static", "js", "service-worker.js"),
        mimetype="application/javascript",
        max_age=0,
    )


@app.route("/notifications")
def notifications_feed():
    user = session.get("user")
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    with get_db() as conn:
        notifications = get_active_notifications(
            conn,
            normalize_user_role(user.get("role", "user")),
            user.get("name", "Unknown"),
        )
    payload = [
        {
            "id": row["id"],
            "title": row["title"],
            "message": row["message"],
            "priority": row["priority"],
            "action_url": row["action_url"],
            "created_at": row["created_at"],
            "expires_at": row["expires_at"],
            "audience_role": row["audience_role"],
            "is_read": bool(row["is_read"]),
        }
        for row in notifications
    ]
    return jsonify({"notifications": payload})


@app.route("/notifications/read", methods=["POST"])
def mark_notification_read():
    user = session.get("user")
    if not user:
        return jsonify({"error": "Unauthorized"}), 401
    data = request.get_json(silent=True) or request.form
    notification_id = str(data.get("notification_id", "")).strip()
    if not notification_id.isdigit():
        return jsonify({"error": "Invalid notification id"}), 400
    with get_db() as conn:
        conn.execute(
            "INSERT OR IGNORE INTO notification_reads (notification_id, reader_name, read_at) "
            "VALUES (?, ?, ?)",
            (notification_id, user.get("name", "Unknown"), now_iso()),
        )
        conn.commit()
    return jsonify({"status": "ok"})


@app.route("/customers")
def customers_list():
    query = request.args.get("q", "").strip()
    where_clause = "WHERE c.name LIKE ?" if query else ""
    params = (f"%{query}%",) if query else ()
    with get_db() as conn:
        rows = conn.execute(
            "SELECT c.*, lo.created_at AS last_order_date, "
            "COALESCE(t.total, 0) AS last_order_total "
            "FROM customers c "
            "LEFT JOIN ("
            "  SELECT customer_id, id, created_at, "
            "         ROW_NUMBER() OVER (PARTITION BY customer_id ORDER BY created_at DESC) AS rn "
            "  FROM orders"
            ") lo ON lo.customer_id = c.id AND lo.rn = 1 "
            "LEFT JOIN ("
            "  SELECT order_id, SUM(qty * price_snapshot) AS total "
            "  FROM order_items GROUP BY order_id"
            ") t ON t.order_id = lo.id "
            f"{where_clause} "
            "ORDER BY c.name",
            params,
        ).fetchall()
    credit_limit = parse_float(get_setting("credit_limit", "0"))
    enriched_customers = []
    for row in rows:
        customer = dict(row)
        customer["over_credit_limit"] = (
            credit_limit > 0 and parse_float(customer["outstanding_balance"]) > credit_limit
        )
        enriched_customers.append(customer)
    return render_template(
        "customers_list.html", customers=enriched_customers, query=query
    )


@app.route("/customers/new", methods=["GET", "POST"])
def customer_new():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        phone = request.form.get("phone", "").strip()
        address = request.form.get("address", "").strip()
        phone_code = request.form.get("phone_code", "").strip()
        email = request.form.get("email", "").strip()
        longitude = request.form.get("longitude", "").strip()
        latitude = request.form.get("latitude", "").strip()
        city = request.form.get("city", "").strip()
        state = request.form.get("state", "").strip()
        country = request.form.get("country", "").strip()
        zip_code = request.form.get("zip_code", "").strip()
        buyer_group = request.form.get("buyer_group", "").strip()
        app_user = 1 if request.form.get("app_user") else 0
        priority_flag = request.form.get("priority_flag", "").strip()
        outstanding_balance = parse_float(request.form.get("outstanding_balance"))
        if not name:
            flash("Customer name is required.", "error")
            return render_template("customer_new.html")
        with get_db() as conn:
            conn.execute(
                "INSERT INTO customers "
                "(name, phone, address, created_at, last_visited_at, phone_code, email, longitude, latitude, "
                "city, state, country, zip_code, buyer_group, app_user, priority_flag, outstanding_balance) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    name,
                    phone,
                    address,
                    now_iso(),
                    None,
                    phone_code,
                    email,
                    longitude,
                    latitude,
                    city,
                    state,
                    country,
                    zip_code,
                    buyer_group,
                    app_user,
                    priority_flag,
                    outstanding_balance,
                ),
            )
            conn.commit()
        flash("Customer created.", "success")
        return redirect(url_for("customers_list"))
    return render_template("customer_new.html")


@app.route("/customers/<int:customer_id>/visit", methods=["POST"])
def customer_visit(customer_id):
    with get_db() as conn:
        conn.execute(
            "UPDATE customers SET last_visited_at = ? WHERE id = ?",
            (now_iso(), customer_id),
        )
        conn.commit()
    flash("Customer marked as visited.", "success")
    return redirect(request.referrer or url_for("customers_list"))


@app.route("/customers/<int:customer_id>/priority", methods=["POST"])
def update_customer_priority(customer_id):
    priority_flag = request.form.get("priority_flag", "").strip()
    outstanding_balance = parse_float(request.form.get("outstanding_balance"))
    with get_db() as conn:
        conn.execute(
            "UPDATE customers SET priority_flag = ?, outstanding_balance = ? WHERE id = ?",
            (priority_flag or None, outstanding_balance, customer_id),
        )
        conn.commit()
    flash("Customer summary updated.", "success")
    return redirect(request.referrer or url_for("customers_list"))


@app.route("/catalog/<int:customer_id>")
def catalog(customer_id):
    view = request.args.get("view")
    draft_id = request.args.get("draft_id")
    if view in {"grid", "list"}:
        set_setting("product_view", view)

    view_setting = get_setting("product_view", "grid")
    with get_db() as conn:
        customer = conn.execute(
            "SELECT * FROM customers WHERE id = ?", (customer_id,)
        ).fetchone()
        if not customer:
            flash("Customer not found.", "error")
            return redirect(url_for("home"))
        last_order = conn.execute(
            "SELECT id, created_at FROM orders WHERE customer_id = ? "
            "ORDER BY created_at DESC LIMIT 1",
            (customer_id,),
        ).fetchone()
        last_order_total = 0
        if last_order:
            last_order_total = (
                conn.execute(
                    "SELECT COALESCE(SUM(qty * price_snapshot), 0) AS total "
                    "FROM order_items WHERE order_id = ?",
                    (last_order["id"],),
                ).fetchone()["total"]
                or 0
            )
        ensure_reserved_stock(conn)
        cached = get_cached_catalog_data()
        products = cached["products"]
        categories = cached["categories"]
        subcategories = cached["subcategories"]
        category_images = cached["category_images"]
        subcategory_map = cached["subcategory_map"]
        subcategory_images = cached["subcategory_images"]
        sub_subcategory_map = cached["sub_subcategory_map"]
        sub_subcategory_images = cached["sub_subcategory_images"]
        salesman_id = session.get("user", {}).get("id")
        price_overrides = get_salesman_price_overrides(conn, salesman_id)
        products_json = []
        for product in products:
            product_data = dict(product)
            sku = product_data.get("sku")
            price = product_data.get("price")
            price_credit = product_data.get("price_credit")
            price_cash = product_data.get("price_cash")
            override_price = price_overrides.get(sku)
            has_override = override_price is not None
            if has_override:
                price = override_price
                price_credit = override_price
                if price_cash is not None and price_cash > override_price:
                    price_cash = override_price
            products_json.append(
                {
                    "sku": sku,
                    "title": product_data.get("title"),
                    "category": product_data.get("category"),
                    "subcategory": product_data.get("subcategory"),
                    "sub_subcategory": product_data.get("sub_subcategory"),
                    "image_url_1": product_data.get("image_url_1"),
                    "image_url_2": product_data.get("image_url_2"),
                    "image_url_3": product_data.get("image_url_3"),
                    "image_url_4": product_data.get("image_url_4"),
                    "image_url": product_data.get("image_url"),
                    "price": price,
                    "price_cash": price_cash,
                    "price_credit": price_credit,
                    "has_override": has_override,
                    "reserved_qty": product_data.get("reserved_qty"),
                    "available_qty": product_data.get("available_qty"),
                    "stock_qty": product_data.get("stock_qty"),
                    "mfr_no": product_data.get("mfr_no"),
                    "sort_order": product_data.get("sort_order"),
                }
            )
    draft = None
    with get_db() as conn:
        if draft_id:
            draft = conn.execute(
                "SELECT * FROM drafts WHERE id = ? AND customer_id = ?",
                (draft_id, customer_id),
            ).fetchone()
        if not draft:
            draft = conn.execute(
                "SELECT * FROM drafts WHERE customer_id = ? ORDER BY updated_at DESC, id DESC LIMIT 1",
                (customer_id,),
            ).fetchone()

    catalog_settings = {
        "show_stock": get_setting("show_stock_to_customers", "on") == "on",
        "low_stock_threshold": int(get_setting("low_stock_threshold", "5") or 5),
        "currency_symbol": get_setting("currency_symbol", "SAR"),
    }
    return render_template(
        "catalog.html",
        customer=customer,
        credit_warning=get_credit_warning(customer),
        products=products,
        categories=categories,
        subcategories=subcategories,
        category_images=category_images,
        subcategory_map=subcategory_map,
        subcategory_images=subcategory_images,
        sub_subcategory_map=sub_subcategory_map,
        sub_subcategory_images=sub_subcategory_images,
        products_json=products_json,
        view_setting=view_setting,
        draft=draft,
        last_order_date=last_order["created_at"] if last_order else None,
        last_order_total=last_order_total,
        catalog_settings=catalog_settings,
    )




@app.route("/catalog/<int:customer_id>/resume", methods=["POST"])
def catalog_resume(customer_id):
    items_json = request.form.get("items_json", "[]")
    try:
        items = json.loads(items_json)
    except json.JSONDecodeError:
        items = []
    if not isinstance(items, list):
        items = []

    with get_db() as conn:
        existing = conn.execute(
            "SELECT id FROM drafts WHERE customer_id = ? ORDER BY updated_at DESC, id DESC LIMIT 1",
            (customer_id,),
        ).fetchone()
        if existing:
            conn.execute(
                "UPDATE drafts SET items_json = ?, updated_at = ? WHERE id = ?",
                (json.dumps(items), now_iso(), existing["id"]),
            )
            draft_id = existing["id"]
        else:
            conn.execute(
                "INSERT INTO drafts (customer_id, items_json, created_at, updated_at) VALUES (?, ?, ?, ?)",
                (customer_id, json.dumps(items), now_iso(), now_iso()),
            )
            draft_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.commit()

    return redirect(url_for("catalog", customer_id=customer_id, draft_id=draft_id))


@app.route("/place_order", methods=["POST"])
def place_order():
    customer_id = request.form.get("customer_id")
    items_json = request.form.get("items_json", "[]")
    order_remarks = request.form.get("order_remarks", "").strip()
    try:
        items = json.loads(items_json)
    except json.JSONDecodeError:
        items = []

    if not customer_id or not items:
        flash("No items to place.", "error")
        return redirect(url_for("home"))

    submitter_name = session.get("user", {}).get("name", "Unknown")
    salesman_id = session.get("user", {}).get("id")
    min_price_percent = get_min_price_percent()

    with get_db() as conn:
        customer = conn.execute(
            "SELECT * FROM customers WHERE id = ?", (customer_id,)
        ).fetchone()
        if not customer:
            flash("Customer not found.", "error")
            return redirect(url_for("home"))

        reservation_on = get_setting("reservation_mode", "on") == "on"
        allow_oversell = get_setting("allow_oversell", "off") == "on"
        avail_expr = (
            "COALESCE(s.stock_qty, 0) - COALESCE(r.reserved_qty, 0)"
            if reservation_on
            else "COALESCE(s.stock_qty, 0)"
        )
        price_overrides = get_salesman_price_overrides(conn, salesman_id)
        insufficient = []
        price_violations = []
        valid_items = []
        for item in items:
            sku = item.get("sku")
            qty = int(item.get("qty", 0))
            if qty <= 0 or not sku:
                continue
            row = conn.execute(
                "SELECT p.title, p.price, p.price_credit, p.price_cash, "
                f"{avail_expr} AS available_qty "
                "FROM products p "
                "LEFT JOIN stock s ON p.sku = s.sku "
                "LEFT JOIN reserved_stock r ON p.sku = r.sku "
                "WHERE p.sku = ?",
                (sku,),
            ).fetchone()
            if not row:
                insufficient.append(f"{sku} (not found)")
                continue
            available_qty = row["available_qty"] or 0
            if not allow_oversell and qty > available_qty:
                insufficient.append(f"{sku} (available {available_qty})")
                continue
            base_price = resolve_product_base_price(sku, row, price_overrides)
            price_floor = base_price * min_price_percent / 100
            submitted_price = parse_float(item.get("unit_price"), default=base_price)
            # Salesmen may discount down to the configured floor, never above
            # the resolved (standard or salesman-override) price.
            if submitted_price > base_price + 0.01 or submitted_price < price_floor - 0.01:
                price_violations.append(
                    f"{sku} (allowed {price_floor:.2f}-{base_price:.2f})"
                )
                continue
            valid_items.append((sku, qty, row, round(submitted_price, 2)))

        if insufficient:
            flash(
                "Not enough available stock for: " + ", ".join(insufficient),
                "error",
            )
            return redirect(url_for("catalog", customer_id=customer_id))
        if price_violations:
            flash(
                "Unit price outside the allowed range for: " + ", ".join(price_violations),
                "error",
            )
            return redirect(url_for("catalog", customer_id=customer_id))

        cutoff_iso = (datetime.utcnow() - timedelta(minutes=5)).isoformat()
        duplicate_order_number = has_recent_duplicate_order(
            conn,
            customer_id,
            submitter_name,
            valid_items,
            cutoff_iso,
        )
        if duplicate_order_number:
            flash(
                f"A matching order was already submitted recently ({duplicate_order_number}).",
                "error",
            )
            return redirect(url_for("orders_list", view="history"))

        order_number = generate_order_number(conn)
        created_at = now_iso()
        export_status = "exported"
        order_status = "submitted"
        conn.execute(
            "INSERT INTO orders (order_number, customer_id, created_at, export_status, order_status, submitted_by, remarks) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (
                order_number,
                customer_id,
                created_at,
                export_status,
                order_status,
                submitter_name,
                order_remarks,
            ),
        )
        order_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        order_rows = []
        for sku, qty, row, unit_price in valid_items:
            conn.execute(
                "INSERT INTO order_items (order_id, sku, title_snapshot, price_snapshot, qty) "
                "VALUES (?, ?, ?, ?, ?)",
                (order_id, sku, row["title"], unit_price, qty),
            )
            order_rows.append({"sku": sku, "qty": qty, "price": unit_price})
        adjust_reserved_stock(conn, order_id, 1)
        log_audit(
            conn,
            "order_placed",
            "orders",
            {
                "order_number": order_number,
                "customer_id": customer_id,
                "item_count": len(order_rows),
            },
        )
        notification_title = f"New order {order_number}"
        notification_message = (
            f"{customer['name']} submitted an order with {len(order_rows)} items."
        )
        notification_link = url_for("orders_list", view="history")
        for role in ("warehouse", "admin"):
            create_notification(
                conn,
                notification_title,
                notification_message,
                audience_role=role,
                priority="info",
                action_url=notification_link,
            )
        conn.commit()

    export_warning = None
    if order_rows:
        export_fn = getattr(google_sheets, "export_order_rows", None)
        if export_fn:
            success, message = export_fn(
                order_number,
                customer["name"],
                order_rows,
                remarks=order_remarks,
            )
        else:
            success = False
            message = "Export function unavailable."
        if not success:
            export_warning = message
            with get_db() as conn:
                conn.execute(
                    "UPDATE orders SET export_status = 'export_failed' WHERE order_number = ?",
                    (order_number,),
                )
                conn.commit()

    vat_rate = get_vat_rate()
    subtotal = sum(row["price"] * row["qty"] for row in order_rows)
    vat_amount = subtotal * vat_rate
    return render_template(
        "order_success.html",
        order_number=order_number,
        export_warning=export_warning,
        currency_symbol=get_setting("currency_symbol", "SAR"),
        subtotal=subtotal,
        vat_rate=vat_rate,
        vat_amount=vat_amount,
        total_amount=subtotal + vat_amount,
    )


@app.route("/review_order", methods=["POST"])
def review_order():
    customer_id = request.form.get("customer_id")
    items_json = request.form.get("items_json", "[]")
    try:
        items = json.loads(items_json)
    except json.JSONDecodeError:
        items = []

    if not customer_id or not items:
        flash("No items selected.", "error")
        return redirect(url_for("home"))

    salesman_id = session.get("user", {}).get("id")
    min_price_percent = get_min_price_percent()
    with get_db() as conn:
        customer = conn.execute(
            "SELECT * FROM customers WHERE id = ?", (customer_id,)
        ).fetchone()
        if not customer:
            flash("Customer not found.", "error")
            return redirect(url_for("home"))

        price_overrides = get_salesman_price_overrides(conn, salesman_id)
        order_items = []
        for item in items:
            sku = item.get("sku")
            qty = int(item.get("qty", 0))
            if qty <= 0:
                continue
            product = conn.execute(
                "SELECT sku, title, price, price_cash, price_credit, image_url, image_url_1, image_url_2, image_url_3, image_url_4 FROM products WHERE sku = ?",
                (sku,),
            ).fetchone()
            if not product:
                continue
            price_base = resolve_product_base_price(sku, product, price_overrides)
            price_floor = round(price_base * min_price_percent / 100, 2)
            order_items.append(
                {
                    "sku": product["sku"],
                    "title": product["title"],
                    "price_cash": product["price_cash"],
                    "price_base": price_base,
                    "price_floor": price_floor,
                    "qty": qty,
                    "image_url": product["image_url"]
                    or product["image_url_1"]
                    or product["image_url_2"]
                    or product["image_url_3"]
                    or product["image_url_4"],
                }
            )

    if not order_items:
        flash("No valid items to review.", "error")
        return redirect(url_for("catalog", customer_id=customer_id))

    return render_template(
        "review_order.html",
        customer=customer,
        credit_warning=get_credit_warning(customer),
        items=order_items,
        currency_symbol=get_setting("currency_symbol", "SAR"),
        vat_rate=get_vat_rate(),
        can_edit_price=min_price_percent < 100,
    )


@app.route("/save_draft", methods=["POST"])
def save_draft():
    customer_id = request.form.get("customer_id")
    items_json = request.form.get("items_json", "[]")
    if not customer_id:
        return {"success": False, "message": "Customer is required."}, 400
    try:
        items = json.loads(items_json)
    except json.JSONDecodeError:
        return {"success": False, "message": "Invalid draft data."}, 400
    if not items:
        return {"success": False, "message": "No items to save."}, 400

    with get_db() as conn:
        draft = conn.execute(
            "SELECT id FROM drafts WHERE customer_id = ?",
            (customer_id,),
        ).fetchone()
        if draft:
            conn.execute(
                "UPDATE drafts SET items_json = ?, updated_at = ? WHERE id = ?",
                (items_json, now_iso(), draft["id"]),
            )
            draft_id = draft["id"]
        else:
            conn.execute(
                "INSERT INTO drafts (customer_id, items_json, created_at, updated_at) VALUES (?, ?, ?, ?)",
                (customer_id, items_json, now_iso(), now_iso()),
            )
            draft_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.commit()

    return {"success": True, "draft_id": draft_id}


@app.route("/drafts")
def drafts_list():
    with get_db() as conn:
        drafts = conn.execute(
            "SELECT d.id, d.items_json, d.updated_at, c.id AS customer_id, c.name AS customer_name "
            "FROM drafts d JOIN customers c ON d.customer_id = c.id "
            "ORDER BY d.updated_at DESC"
        ).fetchall()
        products = conn.execute(
            "SELECT sku, price_credit, price FROM products"
        ).fetchall()
    price_lookup = {
        row["sku"]: row["price_credit"] if row["price_credit"] is not None else row["price"]
        for row in products
    }
    draft_summaries = []
    for draft in drafts:
        try:
            items = json.loads(draft["items_json"])
        except json.JSONDecodeError:
            items = []
        total_qty = sum(int(item.get("qty", 0)) for item in items)
        total_amount = sum(
            price_lookup.get(item.get("sku"), 0) * int(item.get("qty", 0))
            for item in items
        )
        draft_summaries.append(
            {
                "id": draft["id"],
                "customer_id": draft["customer_id"],
                "customer_name": draft["customer_name"],
                "updated_at": draft["updated_at"],
                "total_qty": total_qty,
                "total_amount": total_amount,
            }
        )
    return render_template("drafts_list.html", drafts=draft_summaries)


@app.route("/drafts/<int:draft_id>/delete", methods=["POST"])
def delete_draft(draft_id):
    with get_db() as conn:
        conn.execute("DELETE FROM drafts WHERE id = ?", (draft_id,))
        conn.commit()
    flash("Draft cleared.", "success")
    return redirect(url_for("drafts_list"))


@app.route("/orders")
def orders_list():
    buyer_query = request.args.get("buyer", "").strip()
    order_query = request.args.get("order_id", "").strip()
    status_filter = request.args.get("status", "").strip()
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    view_filter = get_orders_view_param()
    limit_raw = request.args.get("limit", "50").strip()
    try:
        limit_value = min(max(int(limit_raw), 10), 200)
    except ValueError:
        limit_value = 50

    filters = []
    params = []
    if view_filter == "live":
        filters.append("o.order_status NOT IN ('delivered', 'cancelled')")
    if buyer_query:
        filters.append("c.name LIKE ?")
        params.append(f"%{buyer_query}%")
    if order_query:
        filters.append("o.order_number LIKE ?")
        params.append(f"%{order_query}%")
    if status_filter:
        filters.append("o.order_status = ?")
        params.append(status_filter)
    if start_date:
        filters.append("o.created_at >= ?")
        params.append(f"{start_date}T00:00:00")
    if end_date:
        filters.append("o.created_at <= ?")
        params.append(f"{end_date}T23:59:59")
    where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""

    with get_db() as conn:
        orders = conn.execute(
            "SELECT o.order_number, o.created_at, o.export_status, "
            "COALESCE(o.order_status, 'submitted') AS order_status, "
            "COALESCE(o.submitted_by, 'Unknown') AS submitted_by, "
            "o.assigned_user_id, u.name AS assigned_user_name, "
            "c.name AS customer_name, "
            "COALESCE(SUM(oi.price_snapshot * oi.qty), 0) AS total_amount, "
            "COALESCE(SUM(oi.qty), 0) AS total_qty "
            "FROM orders o "
            "JOIN customers c ON o.customer_id = c.id "
            "LEFT JOIN users u ON o.assigned_user_id = u.id "
            "LEFT JOIN order_items oi ON o.id = oi.order_id "
            f"{where_clause} "
            "GROUP BY o.id "
            "ORDER BY o.created_at DESC "
            "LIMIT ?",
            (*params, limit_value),
        ).fetchall()
        warehouse_users = conn.execute(
            "SELECT id, name FROM users WHERE role IN ('warehouse', 'admin') ORDER BY name"
        ).fetchall()
    return render_template(
        "orders_list.html",
        orders=orders,
        warehouse_users=warehouse_users,
        buyer_query=buyer_query,
        order_query=order_query,
        status_filter=status_filter,
        start_date=start_date,
        end_date=end_date,
        limit_value=limit_value,
        view_filter=view_filter,
    )


@app.route("/dashboard-mobile")
def dashboard_mobile():
    today_label = datetime.utcnow().strftime("%Y-%m-%d")
    with get_db() as conn:
        status_rows = conn.execute(
            "SELECT COALESCE(order_status, 'submitted') AS status, COUNT(*) AS total "
            "FROM orders GROUP BY status"
        ).fetchall()
        status_counts = {row["status"]: row["total"] for row in status_rows}
        todays_orders = conn.execute(
            "SELECT COUNT(*) FROM orders WHERE created_at LIKE ?",
            (f"{today_label}%",),
        ).fetchone()[0]

    tiles = [
        {
            "label": "Today's Order",
            "count": todays_orders,
            "class": "tile-today",
            "icon": "🛒",
            "link": url_for("orders_mobile", start_date=today_label, end_date=today_label),
        },
        {
            "label": "To be processed",
            "count": status_counts.get("submitted", 0),
            "class": "tile-processing",
            "icon": "📦",
            "link": url_for("orders_mobile", status="submitted"),
        },
        {
            "label": "Accepted",
            "count": status_counts.get("accepted", 0),
            "class": "tile-accepted",
            "icon": "✅",
            "link": url_for("orders_mobile", status="accepted"),
        },
        {
            "label": "Shipped",
            "count": status_counts.get("shipped", 0),
            "class": "tile-shipped",
            "icon": "🚚",
            "link": url_for("orders_mobile", status="shipped"),
        },
        {
            "label": "Delivered",
            "count": status_counts.get("delivered", 0),
            "class": "tile-delivered",
            "icon": "📬",
            "link": url_for("orders_mobile", status="delivered"),
        },
        {
            "label": "Cancelled",
            "count": status_counts.get("cancelled", 0),
            "class": "tile-cancelled",
            "icon": "❌",
            "link": url_for("orders_mobile", status="cancelled"),
        },
        {
            "label": "Delayed",
            "count": 0,
            "class": "tile-delayed",
            "icon": "⏱️",
            "link": url_for("orders_mobile", status="delayed"),
        },
        {
            "label": "Rejected",
            "count": 0,
            "class": "tile-rejected",
            "icon": "🚫",
            "link": url_for("orders_mobile", status="rejected"),
        },
    ]
    return render_template(
        "dashboard_mobile.html",
        tiles=tiles,
        today_label=today_label,
    )


@app.route("/orders-mobile")
def orders_mobile():
    status_filter = request.args.get("status", "").strip().lower()
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    sort_dir = "asc" if request.args.get("sort", "").strip().lower() == "asc" else "desc"

    filters = []
    params = []
    if status_filter:
        filters.append("COALESCE(o.order_status, 'submitted') = ?")
        params.append(status_filter)
    if start_date:
        filters.append("o.created_at >= ?")
        params.append(f"{start_date}T00:00:00")
    if end_date:
        filters.append("o.created_at <= ?")
        params.append(f"{end_date}T23:59:59")
    where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""
    order_by_sql = "ASC" if sort_dir == "asc" else "DESC"

    with get_db() as conn:
        orders = conn.execute(
            "SELECT o.order_number, o.created_at, "
            "COALESCE(o.order_status, 'submitted') AS order_status, "
            "c.name AS customer_name, c.phone AS customer_phone, c.email AS customer_email, "
            "COALESCE(SUM(oi.price_snapshot * oi.qty), 0) AS total_amount "
            "FROM orders o "
            "JOIN customers c ON o.customer_id = c.id "
            "LEFT JOIN order_items oi ON o.id = oi.order_id "
            f"{where_clause} "
            "GROUP BY o.id "
            f"ORDER BY o.created_at {order_by_sql} "
            "LIMIT 200",
            params,
        ).fetchall()

    order_cards = []
    for order in orders:
        parsed = parse_iso_date(order["created_at"])
        order_cards.append(
            {
                "order_number": order["order_number"],
                "order_status": order["order_status"],
                "customer_name": order["customer_name"],
                "customer_phone": order["customer_phone"],
                "customer_email": order["customer_email"],
                "total_amount": order["total_amount"],
                "year": parsed.strftime("%Y") if parsed else "",
                "day": parsed.strftime("%d") if parsed else "",
                "month": parsed.strftime("%b") if parsed else "",
            }
        )

    return render_template(
        "orders_mobile.html",
        orders=order_cards,
        status_filter=status_filter,
        start_date=start_date,
        end_date=end_date,
        sort_dir=sort_dir,
    )


@app.route("/orders/<order_number>/download")
def download_order(order_number):
    with get_db() as conn:
        order = conn.execute(
            "SELECT o.order_number, o.created_at, c.name AS customer_name "
            "FROM orders o JOIN customers c ON o.customer_id = c.id "
            "WHERE o.order_number = ?",
            (order_number,),
        ).fetchone()
        if not order:
            flash("Order not found.", "error")
            return redirect(url_for("orders_list"))
        items = conn.execute(
            "SELECT sku, title_snapshot, qty, price_snapshot "
            "FROM order_items WHERE order_id = "
            "(SELECT id FROM orders WHERE order_number = ?)",
            (order_number,),
        ).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Order Number", sanitize_csv_cell(order["order_number"])])
    writer.writerow(["Customer", sanitize_csv_cell(order["customer_name"])])
    writer.writerow(["Created At", order["created_at"]])
    writer.writerow([])
    writer.writerow(["SKU", "Title", "Qty", "Price"])
    for item in items:
        writer.writerow(
            [
                sanitize_csv_cell(item["sku"]),
                sanitize_csv_cell(item["title_snapshot"]),
                item["qty"],
                f"{item['price_snapshot']:.2f}",
            ]
        )
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={order_number}.csv"
        },
    )


def export_orders(order_numbers, as_pdf=False):
    vat_rate = get_vat_rate()
    with get_db() as conn:
        placeholders = ",".join("?" for _ in order_numbers)
        orders = conn.execute(
            "SELECT o.id, o.order_number, o.created_at, c.name AS customer_name, "
            "c.phone AS customer_phone, c.address AS customer_address, c.city AS customer_city, "
            "c.state AS customer_state, c.country AS customer_country "
            f"FROM orders o JOIN customers c ON o.customer_id = c.id "
            f"WHERE o.order_number IN ({placeholders})",
            order_numbers,
        ).fetchall()
        if not orders:
            return Response(
                "No orders found.",
                mimetype="text/plain",
                headers={"Content-Disposition": "attachment; filename=orders.txt"},
            )
        order_map = {
            order["id"]: {"number": order["order_number"], "created_at": order["created_at"]}
            for order in orders
        }
        order_ids = [order["id"] for order in orders]
        items = conn.execute(
            "SELECT oi.order_id, oi.sku, oi.title_snapshot, oi.qty, oi.price_snapshot, "
            "COALESCE(p.image_url, p.image_url_1, p.image_url_2, p.image_url_3, p.image_url_4) "
            "AS image_url "
            "FROM order_items oi "
            "LEFT JOIN products p ON oi.sku = p.sku "
            f"WHERE oi.order_id IN ({','.join('?' for _ in order_ids)})",
            order_ids,
        ).fetchall()

    if as_pdf:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            leftMargin=36,
            rightMargin=36,
            topMargin=36,
            bottomMargin=36,
        )
        styles = getSampleStyleSheet()
        title_style = styles["Title"]
        title_style.fontSize = 18
        title_style.leading = 22
        normal_style = styles["BodyText"]
        normal_style.leading = 14
        small_style = styles["BodyText"].clone("SmallBody")
        small_style.fontSize = 9
        small_style.leading = 11

        def format_customer_details(order):
            lines = [order["customer_name"]]
            address_parts = [
                order["customer_address"],
                order["customer_city"],
                order["customer_state"],
                order["customer_country"],
            ]
            address = ", ".join([part for part in address_parts if part])
            if address:
                lines.append(address)
            if order["customer_phone"]:
                lines.append(f"Phone: {order['customer_phone']}")
            return "<br/>".join(lines)

        def order_image_cell(url, allow_images=True):
            if not allow_images or not url:
                return Paragraph("—", styles["BodyText"])
            try:
                with urllib.request.urlopen(url, timeout=1) as response:
                    data = response.read()
                img = Image(io.BytesIO(data))
                img.drawWidth = 0.5 * inch
                img.drawHeight = 0.5 * inch
                return img
            except Exception:
                return Paragraph("—", styles["BodyText"])

        story = []
        for index, order in enumerate(orders):
            if index:
                story.append(PageBreak())
            story.append(Paragraph("Sales Order", title_style))
            story.append(Spacer(1, 10))

            header_table = Table(
                [
                    [
                        Paragraph("<b>Order Number</b>", normal_style),
                        Paragraph(order["order_number"], normal_style),
                        Paragraph("<b>Order Date</b>", normal_style),
                        Paragraph(order["created_at"], normal_style),
                    ],
                    [
                        Paragraph("<b>Customer</b>", normal_style),
                        Paragraph(format_customer_details(order), normal_style),
                        Paragraph("<b>Status</b>", normal_style),
                        Paragraph("Submitted", normal_style),
                    ],
                ],
                colWidths=[1.1 * inch, 3.3 * inch, 1.1 * inch, 1.8 * inch],
            )
            header_table.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                        ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.grey),
                    ]
                )
            )
            story.append(header_table)
            story.append(Spacer(1, 12))

            order_items = [item for item in items if item["order_id"] == order["id"]]
            allow_images = len(order_items) <= 15
            item_rows = [
                [
                    Paragraph("<b>Image</b>", small_style),
                    Paragraph("<b>Item Code</b>", small_style),
                    Paragraph("<b>Description</b>", small_style),
                    Paragraph("<b>Qty</b>", small_style),
                    Paragraph("<b>Unit Price</b>", small_style),
                    Paragraph("<b>Line Total</b>", small_style),
                ]
            ]

            subtotal = 0
            for item in order_items:
                line_total = float(item["qty"]) * float(item["price_snapshot"])
                subtotal += line_total
                item_rows.append(
                    [
                        order_image_cell(item["image_url"], allow_images),
                        Paragraph(item["sku"], small_style),
                        Paragraph(item["title_snapshot"], small_style),
                        Paragraph(str(item["qty"]), small_style),
                        Paragraph(f"SAR {item['price_snapshot']:.2f}", small_style),
                        Paragraph(f"SAR {line_total:.2f}", small_style),
                    ]
                )

            if not order_items:
                item_rows.append(
                    [
                        Paragraph("—", small_style),
                        Paragraph("No items", small_style),
                        Paragraph("No items", small_style),
                        Paragraph("0", small_style),
                        Paragraph("SAR 0.00", small_style),
                        Paragraph("SAR 0.00", small_style),
                    ]
                )

            items_table = Table(
                item_rows,
                colWidths=[
                    0.8 * inch,
                    1.2 * inch,
                    2.2 * inch,
                    0.7 * inch,
                    1.1 * inch,
                    1.1 * inch,
                ],
            )
            items_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2f7")),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            story.append(items_table)
            story.append(Spacer(1, 12))

            vat_amount = subtotal * vat_rate
            total_amount = subtotal + vat_amount
            totals_table = Table(
                [
                    ["Subtotal (Excl. VAT)", f"SAR {subtotal:.2f}"],
                    [f"VAT ({vat_rate * 100:.0f}%)", f"SAR {vat_amount:.2f}"],
                    ["Total (Incl. VAT)", f"SAR {total_amount:.2f}"],
                ],
                colWidths=[1.6 * inch, 1.4 * inch],
                hAlign="RIGHT",
            )
            totals_table.setStyle(
                TableStyle(
                    [
                        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#0f172a")),
                        ("FONTNAME", (0, 2), (-1, 2), "Helvetica-Bold"),
                        ("LINEBELOW", (0, 0), (-1, 1), 0.5, colors.grey),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            story.append(totals_table)
            story.append(Spacer(1, 12))
            story.append(
                Paragraph(
                    "Prepared by: ____________________&nbsp;&nbsp;&nbsp;&nbsp;Approved by: ____________________",
                    small_style,
                )
            )

        doc.build(story)
        buffer.seek(0)
        return Response(
            buffer.getvalue(),
            mimetype="application/pdf",
            headers={"Content-Disposition": "attachment; filename=orders.pdf"},
        )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Order Number", "Created At", "SKU", "Title", "Qty", "Price"])
    for item in items:
        writer.writerow(
            [
                sanitize_csv_cell(order_map.get(item["order_id"], {}).get("number", "")),
                order_map.get(item["order_id"], {}).get("created_at", ""),
                sanitize_csv_cell(item["sku"]),
                sanitize_csv_cell(item["title_snapshot"]),
                item["qty"],
                f"{item['price_snapshot']:.2f}",
            ]
        )
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=orders.csv"},
    )


def build_quotation_pdf(customer, quote_items, quote_ref, vat_rate, currency_symbol):
    """Build a shareable quotation PDF (with item images) from cart items that
    have not yet been committed as an order."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontSize = 18
    title_style.leading = 22
    normal_style = styles["BodyText"]
    normal_style.leading = 14
    small_style = styles["BodyText"].clone("SmallBody")
    small_style.fontSize = 9
    small_style.leading = 11

    def format_customer_details(customer_row):
        lines = [customer_row["name"]]
        address_parts = [
            customer_row["address"],
            customer_row["city"],
            customer_row["state"],
            customer_row["country"],
        ]
        address = ", ".join([part for part in address_parts if part])
        if address:
            lines.append(address)
        if customer_row["phone"]:
            lines.append(f"Phone: {customer_row['phone']}")
        return "<br/>".join(lines)

    def image_cell(url, allow_images=True):
        if not allow_images or not url:
            return Paragraph("—", styles["BodyText"])
        try:
            with urllib.request.urlopen(url, timeout=1) as response:
                data = response.read()
            img = Image(io.BytesIO(data))
            img.drawWidth = 0.5 * inch
            img.drawHeight = 0.5 * inch
            return img
        except Exception:
            return Paragraph("—", styles["BodyText"])

    today = datetime.utcnow()
    valid_until = (today + timedelta(days=14)).strftime("%d %b %Y")

    story = [Paragraph("Quotation", title_style), Spacer(1, 10)]

    header_table = Table(
        [
            [
                Paragraph("<b>Quotation Ref</b>", normal_style),
                Paragraph(quote_ref, normal_style),
                Paragraph("<b>Date</b>", normal_style),
                Paragraph(today.strftime("%d %b %Y"), normal_style),
            ],
            [
                Paragraph("<b>Customer</b>", normal_style),
                Paragraph(format_customer_details(customer), normal_style),
                Paragraph("<b>Valid Until</b>", normal_style),
                Paragraph(valid_until, normal_style),
            ],
        ],
        colWidths=[1.1 * inch, 3.3 * inch, 1.1 * inch, 1.8 * inch],
    )
    header_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.grey),
            ]
        )
    )
    story.append(header_table)
    story.append(Spacer(1, 12))

    allow_images = len(quote_items) <= 15
    item_rows = [
        [
            Paragraph("<b>Image</b>", small_style),
            Paragraph("<b>Item Code</b>", small_style),
            Paragraph("<b>Description</b>", small_style),
            Paragraph("<b>Qty</b>", small_style),
            Paragraph("<b>Unit Price</b>", small_style),
            Paragraph("<b>Line Total</b>", small_style),
        ]
    ]

    subtotal = 0.0
    for item in quote_items:
        line_total = float(item["qty"]) * float(item["unit_price"])
        subtotal += line_total
        item_rows.append(
            [
                image_cell(item["image_url"], allow_images),
                Paragraph(item["sku"], small_style),
                Paragraph(item["title"], small_style),
                Paragraph(str(item["qty"]), small_style),
                Paragraph(f"{currency_symbol} {item['unit_price']:.2f}", small_style),
                Paragraph(f"{currency_symbol} {line_total:.2f}", small_style),
            ]
        )

    items_table = Table(
        item_rows,
        colWidths=[0.8 * inch, 1.2 * inch, 2.2 * inch, 0.7 * inch, 1.1 * inch, 1.1 * inch],
    )
    items_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2f7")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(items_table)
    story.append(Spacer(1, 12))

    vat_amount = subtotal * vat_rate
    total_amount = subtotal + vat_amount
    totals_table = Table(
        [
            ["Subtotal (Excl. VAT)", f"{currency_symbol} {subtotal:.2f}"],
            [f"VAT ({vat_rate * 100:.0f}%)", f"{currency_symbol} {vat_amount:.2f}"],
            ["Total (Incl. VAT)", f"{currency_symbol} {total_amount:.2f}"],
        ],
        colWidths=[1.6 * inch, 1.4 * inch],
        hAlign="RIGHT",
    )
    totals_table.setStyle(
        TableStyle(
            [
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#0f172a")),
                ("FONTNAME", (0, 2), (-1, 2), "Helvetica-Bold"),
                ("LINEBELOW", (0, 0), (-1, 1), 0.5, colors.grey),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(totals_table)
    story.append(Spacer(1, 14))
    story.append(
        Paragraph(
            "This is a price quotation, not a tax invoice or order confirmation. "
            "Prices are subject to stock availability at the time of ordering.",
            small_style,
        )
    )

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


@app.route("/quotation", methods=["POST"])
def quotation():
    customer_id = request.form.get("customer_id")
    items_json = request.form.get("items_json", "[]")
    try:
        items = json.loads(items_json)
    except json.JSONDecodeError:
        items = []

    if not customer_id or not items:
        flash("No items to quote.", "error")
        return redirect(url_for("home"))

    salesman_id = session.get("user", {}).get("id")
    min_price_percent = get_min_price_percent()
    with get_db() as conn:
        customer = conn.execute(
            "SELECT * FROM customers WHERE id = ?", (customer_id,)
        ).fetchone()
        if not customer:
            flash("Customer not found.", "error")
            return redirect(url_for("home"))

        price_overrides = get_salesman_price_overrides(conn, salesman_id)
        price_violations = []
        quote_items = []
        for item in items:
            sku = item.get("sku")
            qty = int(item.get("qty", 0))
            if qty <= 0 or not sku:
                continue
            product = conn.execute(
                "SELECT sku, title, price, price_credit, image_url, image_url_1, "
                "image_url_2, image_url_3, image_url_4 FROM products WHERE sku = ?",
                (sku,),
            ).fetchone()
            if not product:
                continue
            base_price = resolve_product_base_price(sku, product, price_overrides)
            price_floor = base_price * min_price_percent / 100
            unit_price = parse_float(item.get("unit_price"), default=base_price)
            # Reject rather than silently clamp: a quotation should never show
            # a different price than what the salesman asked for, or place_order
            # could later refuse the exact price already shared with the customer.
            if unit_price > base_price + 0.01 or unit_price < price_floor - 0.01:
                price_violations.append(
                    f"{sku} (allowed {price_floor:.2f}-{base_price:.2f})"
                )
                continue
            quote_items.append(
                {
                    "sku": sku,
                    "title": product["title"],
                    "qty": qty,
                    "unit_price": round(unit_price, 2),
                    "image_url": product["image_url"]
                    or product["image_url_1"]
                    or product["image_url_2"]
                    or product["image_url_3"]
                    or product["image_url_4"],
                }
            )

        if price_violations:
            flash(
                "Unit price outside the allowed range for: " + ", ".join(price_violations),
                "error",
            )
            return redirect(url_for("catalog", customer_id=customer_id))
        if not quote_items:
            flash("No valid items to quote.", "error")
            return redirect(url_for("catalog", customer_id=customer_id))

        quote_ref = f"QUO-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}"
        log_audit(
            conn,
            "quotation_generated",
            "quotations",
            {"customer_id": customer_id, "item_count": len(quote_items), "quote_ref": quote_ref},
        )
        conn.commit()

    pdf_bytes = build_quotation_pdf(
        customer,
        quote_items,
        quote_ref,
        get_vat_rate(),
        get_setting("currency_symbol", "SAR"),
    )
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={quote_ref}.pdf"},
    )


@app.route("/orders/<order_number>/status", methods=["POST"])
def update_order_status(order_number):
    warehouse_guard = require_warehouse()
    if warehouse_guard:
        return warehouse_guard
    new_status = request.form.get("order_status", "").strip()
    view_filter = get_orders_view_param()
    allowed = {"submitted", "accepted", "packed", "shipped", "delivered", "cancelled"}
    if new_status not in allowed:
        flash("Invalid status.", "error")
        return redirect(url_for("orders_list", view=view_filter))
    with get_db() as conn:
        ensure_reserved_stock(conn)
        order = conn.execute(
            "SELECT id, order_status FROM orders WHERE order_number = ?",
            (order_number,),
        ).fetchone()
        if not order:
            flash("Order not found.", "error")
            return redirect(url_for("orders_list", view=view_filter))
        reserved_statuses = {"submitted", "accepted", "packed", "shipped"}
        was_reserved = order["order_status"] in reserved_statuses
        will_reserve = new_status in reserved_statuses
        conn.execute(
            "UPDATE orders SET order_status = ? WHERE order_number = ?",
            (new_status, order_number),
        )
        if was_reserved != will_reserve:
            adjust_reserved_stock(conn, order["id"], 1 if will_reserve else -1)
        log_audit(
            conn,
            "order_status_update",
            "orders",
            {
                "order_number": order_number,
                "from_status": order["order_status"],
                "to_status": new_status,
            },
        )
        conn.commit()
    flash(f"Order {order_number} updated.", "success")
    return redirect(url_for("orders_list", view=view_filter))


@app.route("/orders/<order_number>/assign", methods=["POST"])
def assign_order(order_number):
    warehouse_guard = require_warehouse()
    if warehouse_guard:
        return warehouse_guard
    assignee_id = request.form.get("assigned_user_id")
    view_filter = get_orders_view_param()
    with get_db() as conn:
        conn.execute(
            "UPDATE orders SET assigned_user_id = ? WHERE order_number = ?",
            (assignee_id or None, order_number),
        )
        log_audit(
            conn,
            "order_assign",
            "orders",
            {"order_number": order_number, "assigned_user_id": assignee_id or None},
        )
        conn.commit()
    flash(f"Order {order_number} assigned.", "success")
    return redirect(url_for("orders_list", view=view_filter))


@app.route("/orders/<order_number>/items")
def order_items(order_number):
    with get_db() as conn:
        order = conn.execute(
            "SELECT id, order_number FROM orders WHERE order_number = ?",
            (order_number,),
        ).fetchone()
        if not order:
            return {"error": "Not found"}, 404
        items = conn.execute(
            "SELECT oi.sku, oi.title_snapshot, oi.qty, oi.price_snapshot, "
            "COALESCE(p.image_url, p.image_url_1, p.image_url_2, p.image_url_3, p.image_url_4) AS image_url "
            "FROM order_items oi "
            "LEFT JOIN products p ON oi.sku = p.sku "
            "WHERE oi.order_id = ?",
            (order["id"],),
        ).fetchall()
    return {
        "order_number": order["order_number"],
        "items": [
            {
                "sku": item["sku"],
                "title": item["title_snapshot"],
                "qty": item["qty"],
                "price": item["price_snapshot"],
                "image_url": item["image_url"],
            }
            for item in items
        ],
    }


@app.route("/orders/bulk_action", methods=["POST"])
def orders_bulk_action():
    warehouse_guard = require_warehouse()
    if warehouse_guard:
        return warehouse_guard
    action = request.form.get("bulk_action")
    order_numbers = request.form.get("order_numbers", "")
    order_list = [num for num in order_numbers.split(",") if num]
    view_filter = get_orders_view_param()
    if not order_list:
        flash("Select at least one order.", "error")
        return redirect(url_for("orders_list", view=view_filter))
    if action == "status":
        new_status = request.form.get("bulk_status", "").strip()
        allowed = {"submitted", "accepted", "packed", "shipped", "delivered", "cancelled"}
        if new_status not in allowed:
            flash("Invalid status.", "error")
            return redirect(url_for("orders_list", view=view_filter))
        with get_db() as conn:
            ensure_reserved_stock(conn)
            reserved_statuses = {"submitted", "accepted", "packed", "shipped"}
            orders = conn.execute(
                f"SELECT id, order_status FROM orders WHERE order_number IN ({','.join('?' for _ in order_list)})",
                order_list,
            ).fetchall()
            for order in orders:
                was_reserved = order["order_status"] in reserved_statuses
                will_reserve = new_status in reserved_statuses
                if was_reserved != will_reserve:
                    adjust_reserved_stock(conn, order["id"], 1 if will_reserve else -1)
            conn.execute(
                f"UPDATE orders SET order_status = ? WHERE order_number IN ({','.join('?' for _ in order_list)})",
                (new_status, *order_list),
            )
            log_audit(
                conn,
                "order_bulk_status_update",
                "orders",
                {"order_numbers": order_list, "to_status": new_status},
            )
            conn.commit()
        flash("Orders updated.", "success")
        return redirect(url_for("orders_list", view=view_filter))
    if action in {"export_csv", "export_pdf"}:
        return export_orders(order_list, action == "export_pdf")
    flash("Unsupported action.", "error")
    return redirect(url_for("orders_list", view=view_filter))


@app.route("/admin", methods=["GET", "POST"])
def admin():
    admin_guard = require_admin()
    if admin_guard:
        return admin_guard
    message = None
    warning = None
    if request.method == "POST":
        action = request.form.get("action")
        if action == "upload_catalogue":
            file = request.files.get("catalogue_csv")
            upload_mode = request.form.get("upload_mode", "replace")
            if not file:
                warning = "Please upload a CSV or Excel file."
            else:
                rows, warning = load_catalog_rows(file)
                if not warning:
                    required_columns = {"sku", "title", "price"}
                    missing_columns = required_columns - set(
                        rows[0].keys() if rows else []
                    )
                    if missing_columns:
                        warning = (
                            "Missing required columns: "
                            + ", ".join(sorted(missing_columns))
                        )
                        rows = []
                if not warning:
                    with get_db() as conn:
                        inserted = 0
                        updated = 0
                        skipped = 0
                        replaced_count = 0
                        if upload_mode == "replace":
                            replaced_count = conn.execute(
                                "SELECT COUNT(*) FROM products"
                            ).fetchone()[0]
                            log_audit(
                                conn,
                                "catalogue_replace_delete",
                                "products",
                                {
                                    "products_deleted": replaced_count,
                                    "note": "Full catalogue replace - all products not in the "
                                    "uploaded file were removed.",
                                },
                            )
                            conn.execute("DELETE FROM products")
                        for row in rows:
                            sku = row.get("sku", "").strip()
                            if not sku:
                                skipped += 1
                                continue
                            existing = conn.execute(
                                "SELECT 1 FROM products WHERE sku = ?",
                                (sku,),
                            ).fetchone()
                            title = row.get("title", "").strip()
                            price = parse_float(row.get("price", 0))
                            price_cash = parse_float(
                                row.get("cash_price", row.get("price_cash", price)),
                                default=price,
                            )
                            price_credit = parse_float(
                                row.get("credit_price", row.get("price_credit", price)),
                                default=price,
                            )
                            category = row.get("category", "").strip()
                            subcategory = row.get("subcategory", "").strip()
                            sub_subcategory = row.get("sub_subcategory", "").strip()
                            mfr_no = row.get("mfr_no", "").strip()
                            sort_order = parse_int(row.get("sort_order", ""))
                            image_url = row.get("image_url", "").strip()
                            image_url_1 = row.get("image_url_1", "").strip()
                            image_url_2 = row.get("image_url_2", "").strip()
                            image_url_3 = row.get("image_url_3", "").strip()
                            image_url_4 = row.get("image_url_4", "").strip()
                            conn.execute(
                                "INSERT INTO products (sku, title, price, price_cash, price_credit, category, subcategory, sub_subcategory, mfr_no, sort_order, image_url, image_url_1, image_url_2, image_url_3, image_url_4, is_active) "
                                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1) "
                                "ON CONFLICT(sku) DO UPDATE SET "
                                "title = excluded.title, price = excluded.price, price_cash = excluded.price_cash, "
                                "price_credit = excluded.price_credit, category = excluded.category, "
                                "subcategory = excluded.subcategory, sub_subcategory = excluded.sub_subcategory, "
                                "mfr_no = excluded.mfr_no, sort_order = excluded.sort_order, image_url = excluded.image_url, "
                                "image_url_1 = excluded.image_url_1, image_url_2 = excluded.image_url_2, "
                                "image_url_3 = excluded.image_url_3, image_url_4 = excluded.image_url_4, is_active = 1",
                                (
                                    sku,
                                    title,
                                    price_credit,
                                    price_cash,
                                    price_credit,
                                    category,
                                    subcategory,
                                    sub_subcategory,
                                    mfr_no,
                                    sort_order,
                                    image_url,
                                    image_url_1,
                                    image_url_2,
                                    image_url_3,
                                    image_url_4,
                                ),
                            )
                            if existing:
                                updated += 1
                            else:
                                inserted += 1

                            category_image = row.get("category_image_url", "").strip()
                            if category:
                                conn.execute(
                                    "INSERT INTO categories (name, image_url) VALUES (?, ?) "
                                    "ON CONFLICT(name) DO UPDATE SET image_url = COALESCE(excluded.image_url, categories.image_url)",
                                    (category, category_image or None),
                                )

                            subcategory_image = row.get("subcategory_image_url", "").strip()
                            if subcategory:
                                conn.execute(
                                    "INSERT INTO subcategories (name, category_name, image_url) VALUES (?, ?, ?) "
                                    "ON CONFLICT(name) DO UPDATE SET category_name = excluded.category_name, "
                                    "image_url = COALESCE(excluded.image_url, subcategories.image_url)",
                                    (subcategory, category or None, subcategory_image or None),
                                )
                            sub_subcategory_image = row.get(
                                "sub_subcategory_image_url", ""
                            ).strip()
                            if sub_subcategory:
                                conn.execute(
                                    "INSERT INTO sub_subcategories (name, subcategory_name, image_url) VALUES (?, ?, ?) "
                                    "ON CONFLICT(name) DO UPDATE SET subcategory_name = excluded.subcategory_name, "
                                    "image_url = COALESCE(excluded.image_url, sub_subcategories.image_url)",
                                    (
                                        sub_subcategory,
                                        subcategory or None,
                                        sub_subcategory_image or None,
                                    ),
                                )
                        conn.commit()
                        log_audit(
                            conn,
                            "catalogue_upload",
                            "products",
                            {
                                "mode": upload_mode,
                                "processed": len(rows),
                                "inserted": inserted,
                                "updated": updated,
                                "skipped": skipped,
                            },
                        )
                        conn.commit()
                if not warning:
                    message = (
                        f"Processed {len(rows)} rows. Inserted {inserted}, updated {updated}, skipped {skipped}."
                    )
        elif action == "sync_stock":
            success, msg = google_sheets.sync_stock_from_sheet()
            if success:
                message = msg
                with get_db() as conn:
                    log_audit(conn, "sync_stock", "stock", {"status": "success"})
                    conn.commit()
            else:
                warning = msg
        elif action == "retry_exports":
            success, msg = google_sheets.retry_failed_exports()
            if success:
                message = msg
                with get_db() as conn:
                    log_audit(conn, "retry_exports", "orders", {"status": "success"})
                    conn.commit()
            else:
                warning = msg
        elif action == "add_category_image":
            category_name = request.form.get("category_name", "").strip()
            image_url = request.form.get("category_image_url", "").strip()
            if category_name and image_url:
                with get_db() as conn:
                    conn.execute(
                        "INSERT INTO categories (name, image_url) VALUES (?, ?) "
                        "ON CONFLICT(name) DO UPDATE SET image_url = excluded.image_url",
                        (category_name, image_url),
                    )
                    conn.commit()
                message = "Category image updated."
            else:
                warning = "Category name and image URL are required."
        elif action == "add_subcategory_image":
            subcategory_name = request.form.get("subcategory_name", "").strip()
            category_name = request.form.get("subcategory_category", "").strip()
            image_url = request.form.get("subcategory_image_url", "").strip()
            if subcategory_name and image_url:
                with get_db() as conn:
                    conn.execute(
                        "INSERT INTO subcategories (name, category_name, image_url) VALUES (?, ?, ?) "
                        "ON CONFLICT(name) DO UPDATE SET category_name = excluded.category_name, image_url = excluded.image_url",
                        (subcategory_name, category_name or None, image_url),
                    )
                    conn.commit()
                message = "Subcategory image updated."
            else:
                warning = "Subcategory name and image URL are required."
        elif action == "add_sub_subcategory_image":
            sub_subcategory_name = request.form.get("sub_subcategory_name", "").strip()
            parent_subcategory = request.form.get("sub_subcategory_parent", "").strip()
            image_url = request.form.get("sub_subcategory_image_url", "").strip()
            if sub_subcategory_name and image_url:
                with get_db() as conn:
                    conn.execute(
                        "INSERT INTO sub_subcategories (name, subcategory_name, image_url) VALUES (?, ?, ?) "
                        "ON CONFLICT(name) DO UPDATE SET subcategory_name = excluded.subcategory_name, image_url = excluded.image_url",
                        (sub_subcategory_name, parent_subcategory or None, image_url),
                    )
                    conn.commit()
                message = "Sub-subcategory image updated."
            else:
                warning = "Sub-subcategory name and image URL are required."
        elif action == "delete_category_image":
            category_name = request.form.get("category_name", "").strip()
            if category_name:
                with get_db() as conn:
                    conn.execute(
                        "UPDATE categories SET image_url = NULL WHERE name = ?",
                        (category_name,),
                    )
                    conn.commit()
                message = "Category image removed."
            else:
                warning = "Category name is required."
        elif action == "delete_subcategory_image":
            subcategory_name = request.form.get("subcategory_name", "").strip()
            if subcategory_name:
                with get_db() as conn:
                    conn.execute(
                        "UPDATE subcategories SET image_url = NULL WHERE name = ?",
                        (subcategory_name,),
                    )
                    conn.commit()
                message = "Subcategory image removed."
            else:
                warning = "Subcategory name is required."
        elif action == "delete_sub_subcategory_image":
            sub_subcategory_name = request.form.get("sub_subcategory_name", "").strip()
            if sub_subcategory_name:
                with get_db() as conn:
                    conn.execute(
                        "UPDATE sub_subcategories SET image_url = NULL WHERE name = ?",
                        (sub_subcategory_name,),
                    )
                    conn.commit()
                message = "Sub-subcategory image removed."
            else:
                warning = "Sub-subcategory name is required."
        elif action == "add_customer":
            name = request.form.get("name", "").strip()
            phone = request.form.get("phone", "").strip()
            address = request.form.get("address", "").strip()
            phone_code = request.form.get("phone_code", "").strip()
            email = request.form.get("email", "").strip()
            longitude = request.form.get("longitude", "").strip()
            latitude = request.form.get("latitude", "").strip()
            city = request.form.get("city", "").strip()
            state = request.form.get("state", "").strip()
            country = request.form.get("country", "").strip()
            zip_code = request.form.get("zip_code", "").strip()
            buyer_group = request.form.get("buyer_group", "").strip()
            app_user = 1 if request.form.get("app_user") else 0
            priority_flag = request.form.get("priority_flag", "").strip()
            outstanding_balance = parse_float(request.form.get("outstanding_balance"))
            if not name:
                warning = "Customer name is required."
            elif not phone:
                warning = "Customer phone number is required."
            elif not address:
                warning = "Customer address is required."
            elif not city:
                warning = "Customer city is required."
            elif not country:
                warning = "Customer country is required."
            else:
                with get_db() as conn:
                    conn.execute(
                        "INSERT INTO customers "
                        "(name, phone, address, created_at, last_visited_at, phone_code, email, longitude, latitude, "
                        "city, state, country, zip_code, buyer_group, app_user, priority_flag, outstanding_balance) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (
                            name,
                            phone,
                            address,
                            now_iso(),
                            None,
                            phone_code,
                            email,
                            longitude,
                            latitude,
                            city,
                            state,
                            country,
                            zip_code,
                            buyer_group,
                            app_user,
                            priority_flag or None,
                            outstanding_balance,
                        ),
                    )
                    log_audit(
                        conn,
                        "customer_add",
                        "customers",
                        {"name": name, "phone": phone},
                    )
                    conn.commit()
                message = "Customer added."
        elif action == "update_customer":
            customer_id = request.form.get("customer_id")
            name = request.form.get("name", "").strip()
            phone = request.form.get("phone", "").strip()
            address = request.form.get("address", "").strip()
            phone_code = request.form.get("phone_code", "").strip()
            email = request.form.get("email", "").strip()
            longitude = request.form.get("longitude", "").strip()
            latitude = request.form.get("latitude", "").strip()
            city = request.form.get("city", "").strip()
            state = request.form.get("state", "").strip()
            country = request.form.get("country", "").strip()
            zip_code = request.form.get("zip_code", "").strip()
            buyer_group = request.form.get("buyer_group", "").strip()
            app_user = 1 if request.form.get("app_user") else 0
            priority_flag = request.form.get("priority_flag", "").strip()
            outstanding_balance = parse_float(request.form.get("outstanding_balance"))
            if not customer_id:
                warning = "Customer not found."
            elif not name:
                warning = "Customer name is required."
            elif not phone:
                warning = "Customer phone number is required."
            elif not address:
                warning = "Customer address is required."
            elif not city:
                warning = "Customer city is required."
            elif not country:
                warning = "Customer country is required."
            else:
                with get_db() as conn:
                    conn.execute(
                        "UPDATE customers SET name = ?, phone = ?, address = ?, phone_code = ?, email = ?, "
                        "longitude = ?, latitude = ?, city = ?, state = ?, country = ?, zip_code = ?, "
                        "buyer_group = ?, app_user = ?, priority_flag = ?, outstanding_balance = ? "
                        "WHERE id = ?",
                        (
                            name,
                            phone,
                            address,
                            phone_code,
                            email,
                            longitude,
                            latitude,
                            city,
                            state,
                            country,
                            zip_code,
                            buyer_group,
                            app_user,
                            priority_flag or None,
                            outstanding_balance,
                            customer_id,
                        ),
                    )
                    log_audit(
                        conn,
                        "customer_update",
                        "customers",
                        {"customer_id": customer_id, "name": name},
                    )
                    conn.commit()
                message = "Customer updated."
        elif action == "delete_customer":
            customer_id = request.form.get("customer_id")
            if not customer_id:
                warning = "Customer not found."
            else:
                with get_db() as conn:
                    existing_customer = conn.execute(
                        "SELECT id, name FROM customers WHERE id = ?",
                        (customer_id,),
                    ).fetchone()
                    if not existing_customer:
                        warning = "Customer not found."
                    else:
                        conn.execute(
                            "DELETE FROM customers WHERE id = ?",
                            (customer_id,),
                        )
                        log_audit(
                            conn,
                            "customer_delete",
                            "customers",
                            {"customer_id": customer_id, "name": existing_customer["name"]},
                        )
                        conn.commit()
                        message = "Customer deleted."
        elif action == "create_notification":
            title = request.form.get("notification_title", "").strip()
            message_body = request.form.get("notification_message", "").strip()
            audience_role = request.form.get("notification_role", "all").strip() or "all"
            priority = request.form.get("notification_priority", "info").strip() or "info"
            action_url = request.form.get("notification_action_url", "").strip()
            expires_at_raw = request.form.get("notification_expires_at", "").strip()
            expires_at = parse_datetime_local(expires_at_raw)
            allowed_roles = {"all", "user", "warehouse", "admin"}
            allowed_priorities = {"info", "warning", "urgent"}
            if not title or not message_body:
                warning = "Notification title and message are required."
            elif audience_role not in allowed_roles:
                warning = "Invalid notification audience."
            elif priority not in allowed_priorities:
                warning = "Invalid notification priority."
            elif expires_at_raw and not expires_at:
                warning = "Invalid expiration date."
            else:
                with get_db() as conn:
                    create_notification(
                        conn,
                        title,
                        message_body,
                        audience_role=audience_role,
                        priority=priority,
                        action_url=action_url or None,
                        expires_at=expires_at,
                    )
                    log_audit(
                        conn,
                        "notification_created",
                        "notifications",
                        {
                            "title": title,
                            "audience_role": audience_role,
                            "priority": priority,
                        },
                    )
                    conn.commit()
                message = "Notification sent."
        elif action == "delete_notification":
            notification_id = request.form.get("notification_id", "").strip()
            if not notification_id.isdigit():
                warning = "Invalid notification id."
            else:
                with get_db() as conn:
                    conn.execute(
                        "DELETE FROM notification_reads WHERE notification_id = ?",
                        (notification_id,),
                    )
                    conn.execute(
                        "DELETE FROM notifications WHERE id = ?",
                        (notification_id,),
                    )
                    log_audit(
                        conn,
                        "notification_deleted",
                        "notifications",
                        {"notification_id": notification_id},
                    )
                    conn.commit()
                message = "Notification deleted."
        elif action == "add_user":
            name = request.form.get("user_name", "").strip()
            email = request.form.get("user_email", "").strip()
            phone = request.form.get("user_phone", "").strip()
            role = request.form.get("user_role", "user").strip() or "user"
            password = request.form.get("user_password", "").strip()
            if not name:
                warning = "User name is required."
            elif not email:
                warning = "User email is required."
            elif not password:
                warning = "User password is required."
            else:
                if email and user_email_exists(email):
                    warning = "Email is already in use."
                else:
                    password_hash = generate_password_hash(password)
                    with get_db() as conn:
                        conn.execute(
                            "INSERT INTO users (name, email, phone, role, password, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                            (name, email, phone, role, password_hash, now_iso()),
                        )
                        log_audit(
                            conn,
                            "user_add",
                            "users",
                            {"name": name, "email": email, "role": role},
                        )
                        conn.commit()
                    message = "User added."
        elif action == "delete_user":
            user_id = request.form.get("user_id")
            current_user_name = session.get("user", {}).get("name")
            if not user_id:
                warning = "User not found."
            else:
                with get_db() as conn:
                    existing_user = conn.execute(
                        "SELECT id, name, role FROM users WHERE id = ?",
                        (user_id,),
                    ).fetchone()
                    if not existing_user:
                        warning = "User not found."
                    elif existing_user["name"] == current_user_name:
                        warning = "You cannot delete your own account."
                    elif existing_user["role"] == "admin":
                        warning = "Admin users cannot be deleted."
                    else:
                        conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
                        log_audit(
                            conn,
                            "user_delete",
                            "users",
                            {"user_id": user_id, "name": existing_user["name"]},
                        )
                        conn.commit()
                        message = "User deleted."
        elif action == "update_user_role":
            user_id = request.form.get("user_id")
            role = request.form.get("user_role", "user").strip() or "user"
            if user_id:
                with get_db() as conn:
                    conn.execute(
                        "UPDATE users SET role = ? WHERE id = ?",
                        (role, user_id),
                    )
                    log_audit(
                        conn,
                        "user_role_update",
                        "users",
                        {"user_id": user_id, "role": role},
                    )
                    conn.commit()
                message = "User role updated."
        elif action == "adjust_stock":
            sku = request.form.get("stock_sku", "").strip()
            qty_raw = request.form.get("stock_qty", "").strip()
            try:
                qty = int(qty_raw)
            except ValueError:
                qty = None
            if not sku or qty is None:
                warning = "Provide SKU and a valid quantity."
            else:
                with get_db() as conn:
                    conn.execute(
                        "INSERT INTO stock (sku, stock_qty, last_synced_at) VALUES (?, ?, ?) "
                        "ON CONFLICT(sku) DO UPDATE SET stock_qty = excluded.stock_qty, last_synced_at = excluded.last_synced_at",
                        (sku, qty, now_iso()),
                    )
                    log_audit(
                        conn,
                        "stock_adjust",
                        "stock",
                        {"sku": sku, "stock_qty": qty},
                    )
                    conn.commit()
                message = "Stock updated."
        elif action == "reset_user_password":
            user_id = request.form.get("user_id")
            new_password = request.form.get("new_password", "").strip()
            if not user_id or not new_password:
                warning = "User and new password are required."
            else:
                with get_db() as conn:
                    conn.execute(
                        "UPDATE users SET password = ? WHERE id = ?",
                        (generate_password_hash(new_password), user_id),
                    )
                    log_audit(
                        conn,
                        "user_password_reset",
                        "users",
                        {"user_id": user_id},
                    )
                    conn.commit()
                message = "Password reset."
        elif action == "save_settings":
            reservation_mode = "on" if request.form.get("reservation_mode") == "on" else "off"
            allow_oversell = "on" if request.form.get("allow_oversell") == "on" else "off"
            show_stock = "on" if request.form.get("show_stock_to_customers") == "on" else "off"
            currency_symbol = request.form.get("currency_symbol", "SAR").strip() or "SAR"
            try:
                low_stock_threshold = max(0, int(request.form.get("low_stock_threshold", "5")))
            except ValueError:
                low_stock_threshold = 5
            credit_limit = max(0.0, parse_float(request.form.get("credit_limit", "0")))
            min_price_percent = max(0.0, min(100.0, parse_float(request.form.get("min_price_percent", "100"), default=100.0)))
            vat_rate_percent = max(0.0, min(100.0, parse_float(request.form.get("vat_rate", "15"), default=15.0)))
            with get_db() as conn:
                for key, val in [
                    ("reservation_mode", reservation_mode),
                    ("allow_oversell", allow_oversell),
                    ("show_stock_to_customers", show_stock),
                    ("currency_symbol", currency_symbol[:10]),
                    ("low_stock_threshold", str(low_stock_threshold)),
                    ("credit_limit", str(credit_limit)),
                    ("min_price_percent", str(min_price_percent)),
                    ("vat_rate", str(vat_rate_percent)),
                ]:
                    conn.execute(
                        "INSERT INTO settings (key, value) VALUES (?, ?) "
                        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                        (key, val),
                    )
                log_audit(conn, "settings_update", "settings", {
                    "reservation_mode": reservation_mode,
                    "allow_oversell": allow_oversell,
                    "show_stock_to_customers": show_stock,
                    "currency_symbol": currency_symbol,
                    "low_stock_threshold": low_stock_threshold,
                    "credit_limit": credit_limit,
                    "min_price_percent": min_price_percent,
                    "vat_rate": vat_rate_percent,
                })
                conn.commit()
            message = "Settings saved."
        elif action == "set_salesman_price":
            user_id = parse_int(request.form.get("pricing_user_id"))
            sku = request.form.get("pricing_sku", "").strip()
            price = parse_float(request.form.get("pricing_price"), default=None)
            if not user_id or not sku or price is None or price < 0:
                warning = "Salesman, SKU, and a valid price are required."
            else:
                with get_db() as conn:
                    user_exists = conn.execute(
                        "SELECT 1 FROM users WHERE id = ?", (user_id,)
                    ).fetchone()
                    product_exists = conn.execute(
                        "SELECT 1 FROM products WHERE sku = ?", (sku,)
                    ).fetchone()
                    if not user_exists or not product_exists:
                        warning = "Unknown salesman or SKU."
                    else:
                        conn.execute(
                            "INSERT INTO salesman_prices (user_id, sku, price, updated_at) VALUES (?, ?, ?, ?) "
                            "ON CONFLICT(user_id, sku) DO UPDATE SET price = excluded.price, updated_at = excluded.updated_at",
                            (user_id, sku, price, now_iso()),
                        )
                        log_audit(
                            conn,
                            "salesman_price_set",
                            "salesman_prices",
                            {"user_id": user_id, "sku": sku, "price": price},
                        )
                        conn.commit()
                        message = "Salesman price saved."
        elif action == "clear_salesman_price":
            user_id = parse_int(request.form.get("pricing_user_id"))
            sku = request.form.get("pricing_sku", "").strip()
            if user_id and sku:
                with get_db() as conn:
                    conn.execute(
                        "DELETE FROM salesman_prices WHERE user_id = ? AND sku = ?",
                        (user_id, sku),
                    )
                    log_audit(
                        conn,
                        "salesman_price_cleared",
                        "salesman_prices",
                        {"user_id": user_id, "sku": sku},
                    )
                    conn.commit()
                message = "Salesman price override removed."
        elif action == "toggle_product_visibility":
            sku = request.form.get("sku", "").strip()
            is_active = request.form.get("is_active", "0")
            if sku:
                with get_db() as conn:
                    conn.execute(
                        "UPDATE products SET is_active = ? WHERE sku = ?",
                        (1 if is_active == "1" else 0, sku),
                    )
                    log_audit(
                        conn,
                        "product_visibility",
                        "products",
                        {"sku": sku, "is_active": is_active},
                    )
                    conn.commit()
                message = "Product visibility updated."
        if action:
            _catalog_cache["timestamp"] = 0
            _catalog_cache["data"] = None

    if warning is None:
        creds_path = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
        if not creds_path or not os.path.exists(creds_path):
            warning = "Google service account credentials are missing. Upload .env and share sheets."

    product_query = request.args.get("q", "").strip()
    product_limit = request.args.get("limit", "50").strip()
    product_page = request.args.get("page", "1").strip()
    try:
        product_limit_value = min(max(int(product_limit), 10), 200)
    except ValueError:
        product_limit_value = 50
    try:
        product_page_value = max(int(product_page), 1)
    except ValueError:
        product_page_value = 1
    product_offset = (product_page_value - 1) * product_limit_value

    pricing_salesman_id = parse_int(request.args.get("salesman_id"))
    pricing_query = request.args.get("psq", "").strip()
    pricing_limit = request.args.get("pslimit", "25").strip()
    pricing_page = request.args.get("pspage", "1").strip()
    try:
        pricing_limit_value = min(max(int(pricing_limit), 10), 100)
    except ValueError:
        pricing_limit_value = 25
    try:
        pricing_page_value = max(int(pricing_page), 1)
    except ValueError:
        pricing_page_value = 1
    pricing_offset = (pricing_page_value - 1) * pricing_limit_value

    with get_db() as conn:
        categories = conn.execute("SELECT name, image_url FROM categories ORDER BY name").fetchall()
        subcategories = conn.execute(
            "SELECT name, category_name, image_url FROM subcategories ORDER BY name"
        ).fetchall()
        sub_subcategories = conn.execute(
            "SELECT name, subcategory_name, image_url FROM sub_subcategories ORDER BY name"
        ).fetchall()
        admin_customers = conn.execute(
            "SELECT id, name, phone, address, phone_code, email, longitude, latitude, "
            "city, state, country, zip_code, buyer_group, app_user, priority_flag, outstanding_balance "
            "FROM customers ORDER BY name"
        ).fetchall()
        users = conn.execute("SELECT * FROM users ORDER BY created_at DESC").fetchall()
        salesmen = conn.execute("SELECT id, name, role FROM users ORDER BY name").fetchall()
        if not pricing_salesman_id and salesmen:
            pricing_salesman_id = salesmen[0]["id"]

        pricing_where = "WHERE p.title LIKE ? OR p.sku LIKE ?" if pricing_query else ""
        pricing_params = [f"%{pricing_query}%", f"%{pricing_query}%"] if pricing_query else []
        pricing_total = conn.execute(
            f"SELECT COUNT(*) AS total FROM products p {pricing_where}",
            pricing_params,
        ).fetchone()["total"]
        pricing_products = conn.execute(
            "SELECT p.sku, p.title, p.price, p.price_credit, sp.price AS override_price "
            "FROM products p "
            "LEFT JOIN salesman_prices sp ON sp.sku = p.sku AND sp.user_id = ? "
            f"{pricing_where} "
            "ORDER BY p.title LIMIT ? OFFSET ?",
            (pricing_salesman_id, *pricing_params, pricing_limit_value, pricing_offset),
        ).fetchall()
        pricing_override_count = 0
        if pricing_salesman_id:
            pricing_override_count = conn.execute(
                "SELECT COUNT(*) AS total FROM salesman_prices WHERE user_id = ?",
                (pricing_salesman_id,),
            ).fetchone()["total"]

        product_params = []
        product_where = ""
        if product_query:
            product_where = "WHERE title LIKE ? OR sku LIKE ?"
            like_query = f"%{product_query}%"
            product_params.extend([like_query, like_query])
        total_products = conn.execute(
            f"SELECT COUNT(*) AS total FROM products {product_where}",
            product_params,
        ).fetchone()["total"]
        products = conn.execute(
            "SELECT sku, title, is_active, "
            "COALESCE(image_url, image_url_1, image_url_2, image_url_3, image_url_4) AS image_url "
            f"FROM products {product_where} ORDER BY title LIMIT ? OFFSET ?",
            (*product_params, product_limit_value, product_offset),
        ).fetchall()
        month_start = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        sales_rows = conn.execute(
            "SELECT o.assigned_user_id, u.name AS rep_name, "
            "COUNT(DISTINCT o.id) AS orders_count, "
            "COALESCE(SUM(oi.qty * oi.price_snapshot), 0) AS revenue "
            "FROM orders o "
            "LEFT JOIN users u ON u.id = o.assigned_user_id "
            "LEFT JOIN order_items oi ON oi.order_id = o.id "
            "WHERE o.created_at >= ? "
            "GROUP BY o.assigned_user_id, u.name "
            "ORDER BY revenue DESC",
            (month_start.isoformat(),),
        ).fetchall()
        sales_performance = []
        for row in sales_rows:
            sales_performance.append(
                {
                    "rep_name": row["rep_name"] or "Unassigned",
                    "orders_count": row["orders_count"],
                    "revenue": row["revenue"] or 0,
                }
            )
        visits_this_month = conn.execute(
            "SELECT COUNT(*) AS total FROM customers WHERE last_visited_at >= ?",
            (month_start.isoformat(),),
        ).fetchone()["total"]
        order_count_month = conn.execute(
            "SELECT COUNT(*) AS total FROM orders WHERE created_at >= ?",
            (month_start.isoformat(),),
        ).fetchone()["total"]
        conversion_rate = 0
        if visits_this_month:
            conversion_rate = round((order_count_month / visits_this_month) * 100, 1)
        order_summary = conn.execute(
            "SELECT customer_id, COUNT(*) AS order_count, MAX(created_at) AS last_order_date "
            "FROM orders GROUP BY customer_id"
        ).fetchall()
        order_summary_map = {
            row["customer_id"]: {
                "order_count": row["order_count"],
                "last_order_date": row["last_order_date"],
            }
            for row in order_summary
        }
        customers = conn.execute(
            "SELECT id, name, last_visited_at, outstanding_balance FROM customers"
        ).fetchall()
        customer_health = []
        for customer in customers:
            summary = order_summary_map.get(customer["id"], {})
            last_order_date = summary.get("last_order_date")
            last_order_dt = None
            if last_order_date:
                try:
                    last_order_dt = datetime.fromisoformat(last_order_date)
                except ValueError:
                    last_order_dt = None
            days_since_order = (
                (datetime.utcnow() - last_order_dt).days if last_order_dt else 999
            )
            order_count = summary.get("order_count", 0) or 0
            score = max(0, 100 - days_since_order * 2 + order_count * 5)
            customer_health.append(
                {
                    "name": customer["name"],
                    "score": min(score, 100),
                    "last_order_date": last_order_date,
                    "outstanding_balance": customer["outstanding_balance"] or 0,
                }
            )
        customer_health.sort(key=lambda item: item["score"])
        audit_logs = conn.execute(
            "SELECT actor, action, entity, details, created_at FROM audit_logs "
            "ORDER BY created_at DESC LIMIT 20"
        ).fetchall()
        admin_notifications = conn.execute(
            "SELECT id, title, message, audience_role, priority, action_url, expires_at, created_at "
            "FROM notifications ORDER BY created_at DESC LIMIT 50"
        ).fetchall()

    view_setting = get_setting("product_view", "grid")
    app_settings = {
        "reservation_mode": get_setting("reservation_mode", "on"),
        "allow_oversell": get_setting("allow_oversell", "off"),
        "show_stock_to_customers": get_setting("show_stock_to_customers", "on"),
        "currency_symbol": get_setting("currency_symbol", "SAR"),
        "low_stock_threshold": int(get_setting("low_stock_threshold", "5") or 5),
        "credit_limit": parse_float(get_setting("credit_limit", "0")),
        "min_price_percent": get_min_price_percent(),
        "vat_rate": parse_float(get_setting("vat_rate", "15")),
    }
    return render_template(
        "admin.html",
        message=message,
        warning=warning,
        view_setting=view_setting,
        app_settings=app_settings,
        categories=categories,
        subcategories=subcategories,
        sub_subcategories=sub_subcategories,
        admin_customers=admin_customers,
        users=users,
        products=products,
        product_query=product_query,
        product_limit=product_limit_value,
        product_page=product_page_value,
        product_total=total_products,
        product_has_more=total_products > (product_offset + product_limit_value),
        sales_performance=sales_performance,
        conversion_rate=conversion_rate,
        customer_health=customer_health[:8],
        audit_logs=audit_logs,
        admin_notifications=admin_notifications,
        salesmen=salesmen,
        pricing_salesman_id=pricing_salesman_id,
        pricing_products=pricing_products,
        pricing_query=pricing_query,
        pricing_limit=pricing_limit_value,
        pricing_page=pricing_page_value,
        pricing_total=pricing_total,
        pricing_has_more=pricing_total > (pricing_offset + pricing_limit_value),
        pricing_override_count=pricing_override_count,
    )


@app.route("/admin/catalogue/download")
def download_catalogue():
    admin_guard = require_admin()
    if admin_guard:
        return admin_guard

    output = io.StringIO()
    writer = csv.writer(output)
    header = [
        "sku",
        "title",
        "price",
        "cash_price",
        "credit_price",
        "category",
        "subcategory",
        "sub_subcategory",
        "mfr_no",
        "sort_order",
        "image_url",
        "image_url_1",
        "image_url_2",
        "image_url_3",
        "image_url_4",
        "category_image_url",
        "subcategory_image_url",
        "sub_subcategory_image_url",
    ]
    writer.writerow(header)

    with get_db() as conn:
        category_images = {
            row["name"]: row["image_url"]
            for row in conn.execute("SELECT name, image_url FROM categories").fetchall()
        }
        subcategory_images = {
            row["name"]: row["image_url"]
            for row in conn.execute(
                "SELECT name, image_url FROM subcategories"
            ).fetchall()
        }
        sub_subcategory_images = {
            row["name"]: row["image_url"]
            for row in conn.execute(
                "SELECT name, image_url FROM sub_subcategories"
            ).fetchall()
        }
        products = conn.execute(
            "SELECT sku, title, price, price_cash, price_credit, category, "
            "subcategory, sub_subcategory, mfr_no, sort_order, image_url, "
            "image_url_1, image_url_2, image_url_3, image_url_4 "
            "FROM products ORDER BY COALESCE(sort_order, 999999), title"
        ).fetchall()

    for product in products:
        writer.writerow(
            [
                product["sku"],
                product["title"],
                product["price"],
                product["price_cash"],
                product["price_credit"],
                product["category"],
                product["subcategory"],
                product["sub_subcategory"],
                product["mfr_no"],
                product["sort_order"],
                product["image_url"],
                product["image_url_1"],
                product["image_url_2"],
                product["image_url_3"],
                product["image_url_4"],
                category_images.get(product["category"], ""),
                subcategory_images.get(product["subcategory"], ""),
                sub_subcategory_images.get(product["sub_subcategory"], ""),
            ]
        )

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=catalogue-current.csv"},
    )


@app.route("/admin/catalogue/sample")
def download_catalogue_sample():
    admin_guard = require_admin()
    if admin_guard:
        return admin_guard

    sample_path = os.path.join(os.path.dirname(__file__), "sample_data", "catalogue_sample.csv")
    return send_file(sample_path, as_attachment=True)


@app.route("/admin/items", methods=["GET", "POST"])
def admin_items():
    admin_guard = require_admin()
    if admin_guard:
        return admin_guard

    message = None
    warning = None
    if request.method == "POST":
        action = request.form.get("action")
        if action == "save_item":
            sku = request.form.get("sku", "").strip()
            title = request.form.get("title", "").strip()
            if not sku or not title:
                warning = "SKU and title are required."
            else:
                price = float(request.form.get("price", 0) or 0)
                price_cash = float(request.form.get("price_cash", 0) or 0)
                price_credit = float(request.form.get("price_credit", 0) or 0)
                category = request.form.get("category", "").strip()
                subcategory = request.form.get("subcategory", "").strip()
                sub_subcategory = request.form.get("sub_subcategory", "").strip()
                mfr_no = request.form.get("mfr_no", "").strip()
                sort_order_raw = request.form.get("sort_order", "").strip()
                try:
                    sort_order = int(sort_order_raw) if sort_order_raw else None
                except ValueError:
                    sort_order = None
                image_url = request.form.get("image_url", "").strip()
                image_url_1 = request.form.get("image_url_1", "").strip()
                image_url_2 = request.form.get("image_url_2", "").strip()
                image_url_3 = request.form.get("image_url_3", "").strip()
                image_url_4 = request.form.get("image_url_4", "").strip()
                is_active = 1 if request.form.get("is_active") == "1" else 0
                with get_db() as conn:
                    conn.execute(
                        "INSERT INTO products (sku, title, price, price_cash, price_credit, category, "
                        "subcategory, sub_subcategory, mfr_no, sort_order, image_url, image_url_1, "
                        "image_url_2, image_url_3, image_url_4, is_active) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) "
                        "ON CONFLICT(sku) DO UPDATE SET "
                        "title = excluded.title, price = excluded.price, price_cash = excluded.price_cash, "
                        "price_credit = excluded.price_credit, category = excluded.category, "
                        "subcategory = excluded.subcategory, sub_subcategory = excluded.sub_subcategory, "
                        "mfr_no = excluded.mfr_no, sort_order = excluded.sort_order, image_url = excluded.image_url, "
                        "image_url_1 = excluded.image_url_1, image_url_2 = excluded.image_url_2, "
                        "image_url_3 = excluded.image_url_3, image_url_4 = excluded.image_url_4, "
                        "is_active = excluded.is_active",
                        (
                            sku,
                            title,
                            price,
                            price_cash,
                            price_credit,
                            category,
                            subcategory,
                            sub_subcategory,
                            mfr_no,
                            sort_order,
                            image_url,
                            image_url_1,
                            image_url_2,
                            image_url_3,
                            image_url_4,
                            is_active,
                        ),
                    )
                    conn.commit()
                message = "Item saved."
        elif action == "delete_item":
            sku = request.form.get("sku", "").strip()
            if not sku:
                warning = "SKU is required."
            else:
                with get_db() as conn:
                    conn.execute("UPDATE products SET is_active = 0 WHERE sku = ?", (sku,))
                    conn.commit()
                message = "Item hidden."
        elif action == "add_category_image":
            category_name = request.form.get("category_name", "").strip()
            image_url = request.form.get("category_image_url", "").strip()
            if category_name and image_url:
                with get_db() as conn:
                    conn.execute(
                        "INSERT INTO categories (name, image_url) VALUES (?, ?) "
                        "ON CONFLICT(name) DO UPDATE SET image_url = excluded.image_url",
                        (category_name, image_url),
                    )
                    log_audit(
                        conn,
                        "category_image_update",
                        "categories",
                        {"category": category_name},
                    )
                    conn.commit()
                message = "Category image updated."
            else:
                warning = "Category name and image URL are required."
        elif action == "delete_category_image":
            category_name = request.form.get("category_name", "").strip()
            if category_name:
                with get_db() as conn:
                    conn.execute(
                        "UPDATE categories SET image_url = NULL WHERE name = ?",
                        (category_name,),
                    )
                    log_audit(
                        conn,
                        "category_image_delete",
                        "categories",
                        {"category": category_name},
                    )
                    conn.commit()
                message = "Category image removed."
            else:
                warning = "Category name is required."
        elif action == "add_subcategory_image":
            subcategory_name = request.form.get("subcategory_name", "").strip()
            category_name = request.form.get("subcategory_category", "").strip()
            image_url = request.form.get("subcategory_image_url", "").strip()
            if subcategory_name and image_url:
                with get_db() as conn:
                    conn.execute(
                        "INSERT INTO subcategories (name, category_name, image_url) VALUES (?, ?, ?) "
                        "ON CONFLICT(name) DO UPDATE SET category_name = excluded.category_name, image_url = excluded.image_url",
                        (subcategory_name, category_name or None, image_url),
                    )
                    log_audit(
                        conn,
                        "subcategory_image_update",
                        "subcategories",
                        {"subcategory": subcategory_name},
                    )
                    conn.commit()
                message = "Subcategory image updated."
            else:
                warning = "Subcategory name and image URL are required."
        elif action == "delete_subcategory_image":
            subcategory_name = request.form.get("subcategory_name", "").strip()
            if subcategory_name:
                with get_db() as conn:
                    conn.execute(
                        "UPDATE subcategories SET image_url = NULL WHERE name = ?",
                        (subcategory_name,),
                    )
                    log_audit(
                        conn,
                        "subcategory_image_delete",
                        "subcategories",
                        {"subcategory": subcategory_name},
                    )
                    conn.commit()
                message = "Subcategory image removed."
            else:
                warning = "Subcategory name is required."
        elif action == "add_sub_subcategory_image":
            sub_subcategory_name = request.form.get("sub_subcategory_name", "").strip()
            parent_subcategory = request.form.get("sub_subcategory_parent", "").strip()
            image_url = request.form.get("sub_subcategory_image_url", "").strip()
            if sub_subcategory_name and image_url:
                with get_db() as conn:
                    conn.execute(
                        "INSERT INTO sub_subcategories (name, subcategory_name, image_url) VALUES (?, ?, ?) "
                        "ON CONFLICT(name) DO UPDATE SET subcategory_name = excluded.subcategory_name, image_url = excluded.image_url",
                        (sub_subcategory_name, parent_subcategory or None, image_url),
                    )
                    log_audit(
                        conn,
                        "sub_subcategory_image_update",
                        "sub_subcategories",
                        {"sub_subcategory": sub_subcategory_name},
                    )
                    conn.commit()
                message = "Sub-subcategory image updated."
            else:
                warning = "Sub-subcategory name and image URL are required."
        elif action == "delete_sub_subcategory_image":
            sub_subcategory_name = request.form.get("sub_subcategory_name", "").strip()
            if sub_subcategory_name:
                with get_db() as conn:
                    conn.execute(
                        "UPDATE sub_subcategories SET image_url = NULL WHERE name = ?",
                        (sub_subcategory_name,),
                    )
                    log_audit(
                        conn,
                        "sub_subcategory_image_delete",
                        "sub_subcategories",
                        {"sub_subcategory": sub_subcategory_name},
                    )
                    conn.commit()
                message = "Sub-subcategory image removed."
            else:
                warning = "Sub-subcategory name is required."
        if action:
            _catalog_cache["timestamp"] = 0
            _catalog_cache["data"] = None

    product_query = request.args.get("q", "").strip()
    product_limit = request.args.get("limit", "25").strip()
    product_page = request.args.get("page", "1").strip()
    edit_sku = request.args.get("sku", "").strip()
    try:
        product_limit_value = min(max(int(product_limit), 10), 100)
    except ValueError:
        product_limit_value = 25
    try:
        product_page_value = max(int(product_page), 1)
    except ValueError:
        product_page_value = 1
    product_offset = (product_page_value - 1) * product_limit_value

    with get_db() as conn:
        categories = conn.execute(
            "SELECT name, MAX(image_url) AS image_url "
            "FROM ("
            "SELECT name, image_url FROM categories "
            "UNION ALL "
            "SELECT DISTINCT category AS name, NULL AS image_url FROM products WHERE category <> ''"
            ") taxonomy "
            "GROUP BY name ORDER BY name"
        ).fetchall()
        subcategories = conn.execute(
            "SELECT name, MAX(category_name) AS category_name, MAX(image_url) AS image_url "
            "FROM ("
            "SELECT name, category_name, image_url FROM subcategories "
            "UNION ALL "
            "SELECT DISTINCT p.subcategory AS name, p.category AS category_name, NULL AS image_url "
            "FROM products p WHERE p.subcategory <> ''"
            ") taxonomy "
            "GROUP BY name ORDER BY name"
        ).fetchall()
        sub_subcategories = conn.execute(
            "SELECT name, MAX(subcategory_name) AS subcategory_name, MAX(image_url) AS image_url "
            "FROM ("
            "SELECT name, subcategory_name, image_url FROM sub_subcategories "
            "UNION ALL "
            "SELECT DISTINCT p.sub_subcategory AS name, p.subcategory AS subcategory_name, NULL AS image_url "
            "FROM products p WHERE p.sub_subcategory <> ''"
            ") taxonomy "
            "GROUP BY name ORDER BY name"
        ).fetchall()
        product_params = []
        product_where = ""
        if product_query:
            product_where = "WHERE title LIKE ? OR sku LIKE ?"
            like_query = f"%{product_query}%"
            product_params.extend([like_query, like_query])
        total_products = conn.execute(
            f"SELECT COUNT(*) AS total FROM products {product_where}",
            product_params,
        ).fetchone()["total"]
        products = conn.execute(
            "SELECT sku, title, price, is_active, "
            "COALESCE(image_url, image_url_1, image_url_2, image_url_3, image_url_4) AS image_url "
            f"FROM products {product_where} ORDER BY title LIMIT ? OFFSET ?",
            (*product_params, product_limit_value, product_offset),
        ).fetchall()
        edit_product = None
        if edit_sku:
            edit_product = conn.execute(
                "SELECT * FROM products WHERE sku = ?",
                (edit_sku,),
            ).fetchone()

    return render_template(
        "admin_items.html",
        message=message,
        warning=warning,
        categories=categories,
        subcategories=subcategories,
        sub_subcategories=sub_subcategories,
        products=products,
        product_query=product_query,
        product_limit=product_limit_value,
        product_page=product_page_value,
        product_total=total_products,
        product_has_more=total_products > (product_offset + product_limit_value),
        edit_product=edit_product,
    )


@app.after_request
def set_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    return response


@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404


@app.errorhandler(500)
def server_error(e):
    return render_template("500.html"), 500


if __name__ == "__main__":
    init_db()
    port = int(os.getenv("PORT", 8000))
    debug = os.getenv("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
